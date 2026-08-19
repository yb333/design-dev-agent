"""artifact_patcher 测试：xlsx / yml 严格 patch（docs/specs/opt/07）。

构造最小制品（术加 xlsx + 代码仓 yml 组），验证：行追加/单元格替换/稳定标识定位/
重复不覆盖（严格 patch）/未知列不动。
"""
import json
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook, load_workbook

from artifact_patcher import patch_xlsx, patch_yml_group, _change_index, main

FIXTURES = Path(__file__).resolve().parent / "fixtures" / "opt"


@pytest.fixture(scope="module")
def v2_and_sql(tmp_path_factory):
    from assemble_ts_baseline import build_ts_baseline
    from assemble_ts_opt import apply_decisions
    demo = json.loads((FIXTURES / "baseline_v1_demo_full.json").read_text(encoding="utf-8"))
    b, _ = build_ts_baseline(demo)
    dec = {"change_type": "add_field", "backfill": "pending", "fields": [{
        "field": "channel_name", "target_table": "dwb_trade_order_d",
        "placed_rules": ["R0002"], "intermediate_tables": [],
        "field_type": "VARCHAR(64)", "field_comment": "渠道名称",
        "design_logic": "x", "transform_type": "direct",
        "source": {"table": "dws.dim_channel", "alias": "c", "field": "channel_name"},
        "new_joins": []}]}
    v2 = apply_decisions(b, dec)
    etl = tmp_path_factory.mktemp("etl")
    (etl / "R0002.sql").write_text(
        "SELECT t.order_id, t.cust_id, SUM(t.amount) AS total_amount, c.channel_name "
        "FROM dws.tmp_trade_order t GROUP BY t.order_id, t.cust_id", encoding="utf-8")
    return v2, etl


def make_xlsx(path: Path):
    wb = Workbook()
    ws = wb.active; ws.title = "RULE"
    ws.append(["规则编码", "规则名称", "(生成的)查询语句", "未知列A"])
    ws.append(["R0001", "订单明细", "SELECT old1", "不许动"])
    ws.append(["R0002", "订单汇总", "SELECT old2", "不许动"])
    tf = wb.create_sheet("TargetFields")
    tf.append(["规则编码", "目标字段", "来源字段", "加密方式", "别名", "字段类型", "备注", "未知列B"])
    tf.append(["R0002", "order_id", "t.order_id", "0", "", "VARCHAR(64)", "", "keep"])
    wb.save(path)
    return path


class TestXlsxPatch:
    def test_row_append_and_cell_replace(self, v2_and_sql, tmp_path):
        v2, etl = v2_and_sql
        src = make_xlsx(tmp_path / "pkg.xlsx")
        dst = tmp_path / "patched" / "pkg.xlsx"
        notes = []
        patch_xlsx(src, dst, v2, etl, notes)
        wb = load_workbook(dst)
        tf = wb["TargetFields"]
        rows = [[c.value for c in r] for r in tf.iter_rows(min_row=2)]
        assert rows[0][1] == "order_id" and rows[0][7] == "keep", "存量行与未知列不动"
        assert any(r[0] == "R0002" and r[1] == "channel_name" and r[2] == "c.channel_name"
                   for r in rows), "追加声明字段行"
        rule = wb["RULE"]
        r2 = [r for r in rule.iter_rows(min_row=2) if r[0].value == "R0002"][0]
        assert "channel_name" in r2[2].value, "SQL 单元格替换为新 SQL"
        assert [r for r in rule.iter_rows(min_row=2) if r[0].value == "R0001"][0][3].value == "不许动"
        assert any("追加行" in n for n in notes) and any("替换单元格" in n for n in notes)

    def test_strict_patch_no_overwrite(self, v2_and_sql, tmp_path):
        """存量已存在同名字段（声明漂移场景）→ 跳过不覆盖。"""
        v2, etl = v2_and_sql
        src = make_xlsx(tmp_path / "pkg.xlsx")
        tf = load_workbook(src)["TargetFields"]
        tf.append(["R0002", "channel_name", "历史遗留声明", "0", "", "", "漂移"])
        src.parent.mkdir(exist_ok=True)
        tf.parent.save(src)
        notes = []
        patch_xlsx(src, tmp_path / "p2" / "pkg.xlsx", v2, etl, notes)
        assert any("跳过" in n and "channel_name" in n for n in notes)


class TestYmlPatch:
    def make_group(self, d: Path):
        d.mkdir(parents=True, exist_ok=True)
        (d / "R0001.yml").write_text(yaml.safe_dump({
            "规则编码": "R0001", "(生成的)查询语句": "SELECT old1",
            "额外信息（其他sheet页信息）": {"TargetFields": [
                {"规则编码": "R0001", "目标字段": "order_id", "来源字段": "a.order_id"}]}},
        allow_unicode=True, sort_keys=False), encoding="utf-8")
        (d / "R0002.yml").write_text(yaml.safe_dump({
            "规则编码": "R0002", "(生成的)查询语句": "SELECT old2",
            "额外信息（其他sheet页信息）": {"TargetFields": [
                {"规则编码": "R0002", "目标字段": "order_id", "来源字段": "t.order_id"}]}},
        allow_unicode=True, sort_keys=False), encoding="utf-8")

    def test_patch_group(self, v2_and_sql, tmp_path):
        v2, etl = v2_and_sql
        src = tmp_path / "repo"; self.make_group(src)
        dst = tmp_path / "patched"; notes = []
        patch_yml_group(src, dst, v2, etl, notes)
        r2 = yaml.safe_load((dst / "R0002.yml").read_text(encoding="utf-8"))
        assert "channel_name" in r2["(生成的)查询语句"], "query_sql 替换"
        tfs = r2["额外信息（其他sheet页信息）"]["TargetFields"]
        assert any(t["目标字段"] == "channel_name" for t in tfs)
        assert tfs[0]["目标字段"] == "order_id", "存量条目不动"
        r1 = yaml.safe_load((dst / "R0001.yml").read_text(encoding="utf-8"))
        assert r1["(生成的)查询语句"] == "SELECT old1", "未落位规则不改"
        assert any("round-trip" in n or "注释" in n for n in notes), "已知限制写进说明"


class TestChangeIndex:
    def test_expansion(self, v2_and_sql):
        v2, _ = v2_and_sql
        idx = _change_index(v2)
        assert idx == [("R0002", "channel_name", idx[0][2])]
        assert idx[0][2]["source_field"] == "c.channel_name"
        assert idx[0][2]["field_type"] == "VARCHAR(64)"
