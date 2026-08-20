"""平台制品包 exporter 测试。

覆盖：
- platform_config 加载 + 兜底
- execution_tasks.xlsx（RULE/GroupVariables/TargetFields/空sheet）
- schedule_tasks.xlsx（tasks/jobs/taskParams）
- 编码全部留空（关键约束）
"""
import json
from pathlib import Path

import pytest
import openpyxl

# conftest 已把 coding scripts 加入 sys.path
from assemble_export import (
    load_platform_config,
    resolve_config_by_schema,
    build_rule_rows,
    build_group_variables,
    build_target_fields,
    generate_execution_excel,
    generate_schedule_excel,
    validate_code_closure,
    AUDIT_FIELDS,
    RULE_COLUMNS,
    GROUPVARS_COLUMNS,
    TARGETFIELDS_COLUMNS,
    TASKS_COLUMNS,
    JOBS_COLUMNS,
    TASKPARAMS_COLUMNS,
    _RULE_COL,
    _cfg,
    _split_schema_table,
)


# ============================================================
# 测试用 ts.json fixture
# ============================================================

@pytest.fixture
def sample_ts():
    """单规则 + 视图的测试数据"""
    return {
        "meta": {
            "target": {
                "f_table": {"schema": "dws", "table": "dwb_xxx_f", "cn": "XXX宽表"},
                "i_view": {"schema": "dws", "table": "dwb_xxx_i", "cn": "XXX宽表"},
            },
            "schedule": {
                "schedule_type": "daily",
                "cron": "0 30 3 * * ?",
                "exec_params": {"P_CYCLE_ID": {"value_type": "string", "desc": "批次号", "standard": True}},
                "lts_params": [
                    {"lts_var": "V_CYCLE_ID", "etl_param": "P_CYCLE_ID", "desc": "批次号"},
                    {"lts_var": "V_GROUP_CODE", "etl_param": "", "desc": "规则组编码"},
                ],
                "tasks": {
                    "f": {
                        "task_name": "task_dwb_xxx_f",
                        "job_name": "Pjob_dwb_xxx_f",
                        "cron": "0 30 3 * * ?",
                        "upstream": [
                            {"table": "ods_order_f", "task": "task_ods_order_f", "dep_type": "宽依赖"},
                            {"table": "dim_product_f", "task": "task_dim_product_f", "dep_type": "宽依赖"},
                        ],
                    },
                    "view": {
                        "task_name": "task_dwb_xxx_i",
                        "job_name": "Pjob_dwb_xxx_i",
                        "cron": "0 30 3 * * ?",
                        "upstream": [{"table": "dwb_xxx_f", "task": "task_dwb_xxx_f", "dep_type": "宽依赖"}],
                    },
                    "dq": {
                        "task_name": "task_dwb_xxx_f_dq",
                        "job_name": "Pjob_dwb_xxx_f_dq",
                        "cron": "0 30 3 * * ?",
                        "upstream": [{"table": "dwb_xxx_i", "task": "task_dwb_xxx_i", "dep_type": "宽依赖"}],
                    },
                },
            },
        },
        "rules": {
            "R0001": {
                "rule_name": "XXX汇总",
                "target_table": "dwb_xxx_f",
                "exec_sequence": 1,
                "is_view_step": False,
                "design_intent": "以订单事实表为主表左关联用户表装配宽表",
                "source_tables": [{"schema": "ods", "table": "ods_order_f", "alias": "a"}],
                "fields": [
                    {"target_field": "order_id", "source_fields": [{"table": "ods_order_f", "field": "order_id", "alias": "a"}]},
                    {"target_field": "order_amt", "source_fields": [{"table": "ods_order_f", "field": "amount", "alias": "a"}]},
                    {"target_field": "del_flag", "source_fields": []},       # 审计字段
                    {"target_field": "crt_cycle_id", "source_fields": []},   # 审计字段
                ],
            },
        },
    }


@pytest.fixture
def sample_config():
    """resolve_config_by_schema 返回的结构（含租户块解析结果）"""
    return {
        "shujia": {
            "appid": "APP001",
            "org_abbr": "crm_tenant",
            "project_code": "SRP_ETL",
            "project_cn": "ETL项目",
            "project_en": "ETL_Project",
            "datasource": "SRP_DWS",
            "business_owner": "zhangsan",
        },
        "lts": {
            "project_name": "SRP_DAILY",
            "task_group": "GROUP_SPRD",
        },
    }


@pytest.fixture
def etl_dir(tmp_path, sample_ts):
    """造一个 ETL SQL 文件"""
    d = tmp_path / "etl"
    d.mkdir()
    (d / "R0001.sql").write_text("SELECT 1 AS order_id, 100 AS order_amt", encoding="utf-8")
    return d


# ============================================================
# 配置加载
# ============================================================

class TestLoadPlatformConfig:

    def test_load_config(self, tmp_path):
        cfg_file = tmp_path / "platform_config.json"
        cfg_file.write_text(json.dumps({
            "default": {"shujia": {"project_code": "XXX"}, "lts": {"project_name": "P"}},
        }), encoding="utf-8")
        result = load_platform_config(str(cfg_file))
        assert result["default"]["shujia"]["project_code"] == "XXX"

    def test_missing_file_returns_empty(self):
        assert load_platform_config("/nonexistent/path.json") == {}

    def test_cfg_fallback(self):
        """缺失字段用兜底值"""
        assert _cfg({}, "project_code", "FALLBACK") == "FALLBACK"
        assert _cfg({"project_code": ""}, "project_code", "FALLBACK") == "FALLBACK"
        assert _cfg({"project_code": "SRP"}, "project_code", "FALLBACK") == "SRP"


class TestResolveConfigBySchema:

    def test_schema_mapping_hit(self):
        """schema 在 mappings 里有 → 用 schema 的配置"""
        raw = {
            "default": {"shujia": {"project_code": "DEFAULT"}, "lts": {"project_name": "DEF"}},
            "schema_mappings": {
                "slprd": {"shujia": {"project_code": "SLPRD"}, "lts": {"project_name": "SLPRD_DAILY"}},
            },
        }
        result = resolve_config_by_schema(raw, "slprd")
        assert result["shujia"]["project_code"] == "SLPRD"
        assert result["lts"]["project_name"] == "SLPRD_DAILY"

    def test_schema_miss_use_default(self):
        """schema 在 mappings 里没有 → 用 default"""
        raw = {
            "default": {"shujia": {"project_code": "DEFAULT"}, "lts": {"project_name": "DEF"}},
            "schema_mappings": {},
        }
        result = resolve_config_by_schema(raw, "unknown_schema")
        assert result["shujia"]["project_code"] == "DEFAULT"

    def test_partial_override(self):
        """schema 只配了部分字段，其余用 default 兜底"""
        raw = {
            "default": {"shujia": {"project_code": "DEFAULT", "datasource": "DWS"}, "lts": {"project_name": "DEF"}},
            "schema_mappings": {
                "slprd": {"shujia": {"project_code": "SLPRD"}},  # 只覆盖 project_code
            },
        }
        result = resolve_config_by_schema(raw, "slprd")
        assert result["shujia"]["project_code"] == "SLPRD"
        assert result["shujia"]["datasource"] == "DWS"  # default 兜底

    def test_empty_config(self):
        """空配置返回空结构（appid 由调用方注入）"""
        result = resolve_config_by_schema({}, "slas")
        assert result == {"shujia": {"appid": ""}, "lts": {}}

    def test_tenant_block_overrides(self):
        """★ 租户块：shujia_tenants[appid] 的 org_abbr/datasource 覆盖 schema 级；
        project_cn/business_owner 也住租户块（appid→项目，不随 schema 变）"""
        raw = {
            "default": {"shujia": {"datasource": "SCHEMA_LEVEL_DS", "project_code": "P1"}},
            "shujia_tenants": {
                "APP001": {"org_abbr": "crm_tenant", "datasource": "TENANT_DS",
                           "project_cn": "CRM域", "business_owner": "zhangsan"},
            },
        }
        result = resolve_config_by_schema(raw, "slprd", appid="APP001")
        assert result["shujia"]["org_abbr"] == "crm_tenant"
        assert result["shujia"]["datasource"] == "TENANT_DS"   # 租户级覆盖 schema 级
        assert result["shujia"]["project_code"] == "P1"        # 非租户属性不受影响
        assert result["shujia"]["appid"] == "APP001"
        assert result["shujia"]["project_cn"] == "CRM域"       # 租户块身份全集
        assert result["shujia"]["business_owner"] == "zhangsan"

    def test_tenant_only_config(self):
        """★ 收敛形态：只有 shujia_tenants 一块（example 的最终形态）也能跑通"""
        raw = {
            "shujia_tenants": {
                "APP001": {"org_abbr": "t1", "datasource": "DS1",
                           "project_cn": "域A", "business_owner": "u1"},
            },
        }
        result = resolve_config_by_schema(raw, "dws", appid="APP001")
        assert result["shujia"]["project_cn"] == "域A"
        assert result["lts"] == {}   # lts 无兜底（调度路径以 ts.tasks 为准）

    def test_no_appid_no_tenant_merge(self):
        """没传 appid → 不做租户合并，datasource 走 schema 级"""
        raw = {
            "default": {"shujia": {"datasource": "SCHEMA_LEVEL_DS"}},
            "shujia_tenants": {"APP001": {"org_abbr": "crm_tenant"}},
        }
        result = resolve_config_by_schema(raw, "slprd")
        assert result["shujia"]["datasource"] == "SCHEMA_LEVEL_DS"
        assert "org_abbr" not in result["shujia"]


# ============================================================
# RULE sheet
# ============================================================

class TestBuildRuleRows:

    def test_rule_row_count(self, sample_ts, sample_config, etl_dir):
        """RULE 行数 = 取数规则 + 参数变量（视图不发术加规则行）"""
        rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        # 1 取数 + 1 参数变量 = 2
        assert len(rows) == 2

    def test_view_does_not_create_rule_row(self, sample_ts, sample_config, etl_dir):
        """★ 回归守护：有 i_view 时不产生视图术加规则行（视图 DDL 走 ddl/ 通道部署）"""
        rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        for row in rows:
            sql = row[_RULE_COL["(生成的）查询语句1"]]
            assert "CREATE VIEW" not in sql and "CREATE OR REPLACE VIEW" not in sql
            assert row[_RULE_COL["目标表"]] != sample_ts["meta"]["target"]["i_view"]["table"]

    def test_placeholder_codes(self, sample_ts, sample_config, etl_dir):
        """★ 编码为占位符：规则编码 = ts 规则码 / PV000N；规则组编码 = GR_{组英文名}"""
        rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        etl_row, pv_row = rows[0], rows[1]
        assert etl_row[_RULE_COL["规则编码"]] == "R0001"
        assert etl_row[_RULE_COL["规则组编码"]] == "GR_dwb_xxx_f"
        assert pv_row[_RULE_COL["规则编码"]] == "PV0001"
        assert pv_row[_RULE_COL["规则组编码"]] == "GR_dwb_xxx_f"

    def test_tenant_columns(self, sample_ts, sample_config, etl_dir):
        """★ 租户ID = appid、组织英文简称 = 租户名（shujia_tenants 解析）"""
        rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        for row in rows:
            assert row[_RULE_COL["租户ID"]] == "APP001"
            assert row[_RULE_COL["组织英文简称"]] == "crm_tenant"

    def test_project_only_cn_filled(self, sample_ts, sample_config, etl_dir):
        """★ 项目只填中文名；编码/英文名出厂留空（内网脚本按中文名补齐）。
        sample_config 故意带 project_code/project_en（旧配置兼容）——验证被忽略不进产物。"""
        rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        for row in rows:
            assert row[_RULE_COL["项目中文名"]] == "ETL项目"
            assert row[_RULE_COL["项目编码"]] == ""
            assert row[_RULE_COL["项目英文名"]] == ""
            assert row[_RULE_COL["子项目编码"]] == ""
            assert row[_RULE_COL["子项目中文名"]] == ""
            assert row[_RULE_COL["子项目英文名"]] == ""

    def test_rule_desc_from_design_intent(self, sample_ts, sample_config, etl_dir):
        """规则描述 ← design_intent；备注留空"""
        rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        etl_row = rows[0]
        assert etl_row[_RULE_COL["规则描述"]] == "以订单事实表为主表左关联用户表装配宽表"
        assert etl_row[_RULE_COL["备注"]] == ""

    def test_long_sql_split_columns(self, sample_ts, sample_config, tmp_path):
        """★ 超长 SQL 分列到查询语句1~N，拼接逐字还原"""
        etl_dir = tmp_path / "etl"
        etl_dir.mkdir()
        long_sql = "SELECT " + ", ".join(f"col_{i} AS c{i}" for i in range(3000))  # > 30000
        (etl_dir / "R0001.sql").write_text(long_sql, encoding="utf-8")
        rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        row = rows[0]
        parts = [row[_RULE_COL[f"(生成的）查询语句{n}"]] for n in range(1, 10)]
        non_empty = [p for p in parts if p]
        assert len(non_empty) >= 2                        # 确实分列了
        assert "".join(non_empty) == long_sql             # 拼接逐字还原

    def test_etl_query_in_statement(self, sample_ts, sample_config, etl_dir):
        """取数规则的查询语句列含 SQL 内容"""
        rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        etl_row = rows[0]
        assert "SELECT 1 AS order_id" in etl_row[_RULE_COL["(生成的）查询语句1"]]

    def test_exec_sequence_filled(self, sample_ts, sample_config, etl_dir):
        """执行序列从 ts 透传（决定加工拓扑）"""
        rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        etl_row = rows[0]
        assert etl_row[_RULE_COL["执行序列"]] == "1"  # sample_ts R0001 exec_sequence=1

    def test_param_rule_form(self, sample_ts, sample_config, etl_dir):
        """参数变量规则形态：类型 12、执行序列 -1、查询语句/运行条件留空"""
        rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        pv_row = rows[-1]
        assert pv_row[_RULE_COL["规则类型"]] == "12"
        assert pv_row[_RULE_COL["规则中文名称"]] == "参数变量规则"
        assert pv_row[_RULE_COL["规则英文名称"]] == "Parameter Variable Rule"
        assert pv_row[_RULE_COL["创建方式"]] == "1"
        assert pv_row[_RULE_COL["执行序列"]] == "-1"
        assert pv_row[_RULE_COL["运行条件"]] == ""
        assert pv_row[_RULE_COL["(生成的）查询语句1"]] == ""

    def test_separate_init_group_gets_own_pv_row(self, sample_ts, sample_config, etl_dir):
        """separate init 规则组有自己的参数变量行（每规则组一条，占位码独立）"""
        (etl_dir / "INIT_R0001.sql").write_text(
            "SELECT 1 AS order_id, 100 AS order_amt WHERE dt <= '20260101'", encoding="utf-8"
        )
        sample_ts["init"] = {
            "mode": "derive", "group_mode": "separate",
            "rules": {"INIT_R0001": {
                "rule_name": "XXX汇总(初始化)", "exec_sequence": 1,
                "target_table": "dwb_xxx_f", "is_view_step": False,
                "load_mode": "truncate_table",
            }},
        }
        rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        # 1 取数 + 1 init 取数 + 主组 pv + init 组 pv = 4
        assert len(rows) == 4
        init_row = rows[1]
        assert init_row[_RULE_COL["规则编码"]] == "INIT_R0001"
        assert init_row[_RULE_COL["规则组编码"]] == "GR_dwb_xxx_f_init"
        assert init_row[_RULE_COL["规则组英文名称"]] == "dwb_xxx_f_init"
        pv_rows = [r for r in rows if r[_RULE_COL["规则类型"]] == "12"]
        assert len(pv_rows) == 2
        by_code = {r[_RULE_COL["规则编码"]]: r for r in pv_rows}
        assert set(by_code) == {"PV0001", "PV0002"}
        assert by_code["PV0001"][_RULE_COL["规则组编码"]] == "GR_dwb_xxx_f"
        assert by_code["PV0002"][_RULE_COL["规则组编码"]] == "GR_dwb_xxx_f_init"

    def test_constants_filled(self, sample_ts, sample_config, etl_dir):
        """固定常量正确"""
        rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        etl_row = rows[0]
        assert etl_row[_RULE_COL["数据库类型"]] == "GaussDB"
        assert etl_row[_RULE_COL["删除模式"]] == "1"
        assert etl_row[_RULE_COL["调度类型"]] == "0"

    def test_column_count_is_82(self, sample_ts, sample_config, etl_dir):
        """RULE 每行 82 列"""
        rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        for row in rows:
            assert len(row) == 82


# ============================================================
# GroupVariables
# ============================================================

class TestBuildGroupVariables:

    def test_vars_from_exec_params(self, sample_ts):
        """参数从 exec_params 来"""
        rows = build_group_variables(sample_ts)
        var_names = [r[1] for r in rows]
        assert "P_CYCLE_ID" in var_names

    def test_default_value_from_ts(self):
        """参数默认值从 ts.default_value 读：static 给值，dynamic 留空（平台注入）。"""
        ts = {"meta": {"schedule": {"exec_params": {
            "P_CYCLE_ID": {"default_value": {"type": "dynamic", "expr": "today_ymdhms"}},
            "BIZ_CODE": {"default_value": "STATIC1"},
        }}}}
        rows = build_group_variables(ts)
        vals = {row[1]: row[5] for row in rows}
        assert vals["P_CYCLE_ID"] == ""       # dynamic → 空，平台运行时注入
        assert vals["BIZ_CODE"] == "STATIC1"  # static 裸串 → 给值

    def test_rule_code_is_pv_placeholder(self, sample_ts):
        """★ 规则编码挂参数变量规则行的占位码 PV0001"""
        rows = build_group_variables(sample_ts)
        for row in rows:
            assert row[0] == "PV0001"

    def test_desc_filled(self, sample_ts):
        """描述 ← exec_params.desc"""
        rows = build_group_variables(sample_ts)
        assert rows[0][8] == "批次号"

    def test_gv_per_group_separate_init(self, sample_ts):
        """separate init：每个规则组各挂一份变量（分别指向该组 pv 行占位码）"""
        sample_ts["init"] = {
            "mode": "derive", "group_mode": "separate",
            "rules": {"INIT_R0001": {
                "rule_name": "XXX汇总(初始化)", "exec_sequence": 1,
                "target_table": "dwb_xxx_f", "is_view_step": False,
            }},
        }
        rows = build_group_variables(sample_ts)
        codes = {row[0] for row in rows}
        assert codes == {"PV0001", "PV0002"}

    def test_no_params_empty_rows(self):
        """无 exec_params → 空列表"""
        ts = {"meta": {"schedule": {"exec_params": {}}}}
        assert build_group_variables(ts) == []


# ============================================================
# TargetFields
# ============================================================

class TestBuildTargetFields:

    def test_filter_audit_fields(self, sample_ts):
        """审计字段被过滤"""
        rows = build_target_fields(sample_ts)
        target_fields = [r[1] for r in rows]
        assert "order_id" in target_fields
        assert "del_flag" not in target_fields
        assert "crt_cycle_id" not in target_fields

    def test_source_field_extracted(self, sample_ts):
        """来源字段固定 s.字段 形态（s 是平台标准别名）；无源留空"""
        rows = build_target_fields(sample_ts)
        by_field = {row[1]: row for row in rows}
        assert by_field["order_amt"][2] == "s.amount"
        assert by_field["order_id"][2] == "s.order_id"
        assert by_field["order_amt"][5] == ""  # 别名不填

    def test_no_source_field_empty(self):
        """无源（COUNT(1) 等表达式派生）来源字段留空"""
        ts = {"rules": {"R0001": {"target_table": "dwb_xxx_f", "fields": [
            {"target_field": "record_cnt", "source_fields": []},
        ]}}}
        rows = build_target_fields(ts)
        assert rows[0][2] == ""

    def test_rule_code_is_ts_code(self, sample_ts):
        """★ 规则编码 = ts 规则码（占位，与 RULE 行对应）"""
        rows = build_target_fields(sample_ts)
        for row in rows:
            assert row[0] == "R0001"


# ============================================================
# Excel 生成（端到端）
# ============================================================

class TestGenerateExecutionExcel:

    def test_10_sheets(self, sample_ts, sample_config, etl_dir, tmp_path):
        """execution Excel 有 10 个 sheet"""
        out = tmp_path / "execution_tasks.xlsx"
        generate_execution_excel(sample_ts, sample_config, etl_dir, out)
        wb = openpyxl.load_workbook(out)
        assert len(wb.sheetnames) == 10
        assert "RULE" in wb.sheetnames
        assert "GroupVariables" in wb.sheetnames
        assert "TargetFields" in wb.sheetnames

    def test_empty_sheets_have_headers_only(self, sample_ts, sample_config, etl_dir, tmp_path):
        """空 sheet 只有表头"""
        out = tmp_path / "execution_tasks.xlsx"
        generate_execution_excel(sample_ts, sample_config, etl_dir, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["ModelRelations"]
        assert ws.max_row == 1  # 只有表头行

    def test_rule_sheet_has_data(self, sample_ts, sample_config, etl_dir, tmp_path):
        """RULE sheet 有数据行"""
        out = tmp_path / "execution_tasks.xlsx"
        generate_execution_excel(sample_ts, sample_config, etl_dir, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["RULE"]
        assert ws.max_row == 3  # 表头 + 1 取数 + 1 参数变量（无视图行）

    def test_closure_ok(self, sample_ts, sample_config, etl_dir):
        """★ 三处占位编码闭合：正常数据无问题"""
        rule_rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        gv_rows = build_group_variables(sample_ts)
        tf_rows = build_target_fields(sample_ts)
        assert validate_code_closure(rule_rows, gv_rows, tf_rows) == []

    def test_closure_catches_dangling(self, sample_ts, sample_config, etl_dir):
        """★ 悬挂引用被抓住：GV/TF 引用了 RULE 没有的编码"""
        rule_rows = build_rule_rows(sample_ts, sample_config, etl_dir)
        gv_rows = build_group_variables(sample_ts)
        tf_rows = build_target_fields(sample_ts)
        tf_rows.append(["R9999", "ghost_field", "s.ghost", "0", "", "", "", ""])
        problems = validate_code_closure(rule_rows, gv_rows, tf_rows)
        assert any("R9999" in p for p in problems)

    def test_closure_blocks_generation(self, sample_ts, sample_config, etl_dir, tmp_path, monkeypatch):
        """★ 闭合校验失败阻断生成（fail loud）"""
        import assemble_export
        monkeypatch.setattr(assemble_export, "build_target_fields",
                            lambda ts: [["R9999", "ghost", "s.ghost", "0", "", "", "", ""]])
        out = tmp_path / "execution_tasks.xlsx"
        with pytest.raises(ValueError, match="R9999"):
            generate_execution_excel(sample_ts, sample_config, etl_dir, out)


class TestGenerateScheduleExcel:

    def test_3_sheets(self, sample_ts, sample_config, tmp_path):
        """schedule Excel 有 3 个 sheet"""
        out = tmp_path / "schedule_tasks.xlsx"
        generate_schedule_excel(sample_ts, sample_config, out)
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == ["tasks", "jobs", "taskParams"]

    def test_tasks_has_f_view_dq(self, sample_ts, sample_config, tmp_path):
        """tasks 有 F 表 + 视图 + DQ 三行"""
        out = tmp_path / "schedule_tasks.xlsx"
        generate_schedule_excel(sample_ts, sample_config, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["tasks"]
        assert ws.max_row == 4  # 表头 + F + 视图 + DQ

    def test_jobs_has_upstream_deps(self, sample_ts, sample_config, tmp_path):
        """jobs 含 upstream 依赖行 + view/dq 执行行和依赖"""
        out = tmp_path / "schedule_tasks.xlsx"
        generate_schedule_excel(sample_ts, sample_config, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["jobs"]
        # 表头 + F执行 + 2依赖 + view执行 + view依赖 + dq执行 + dq依赖 = 8
        assert ws.max_row == 8

    def test_cross_project_upstream_uses_dep_project(self, sample_ts, sample_config, tmp_path):
        """跨项目上游依赖：jobs 依赖行的 project/group 用 upstream 项的（不是当前表的）。"""
        # 构造一个跨项目的上游依赖
        ts_cross = json.loads(json.dumps(sample_ts))  # 深拷贝
        ts_cross["meta"]["schedule"]["tasks"]["f"]["upstream"] = [{
            "table": "ods_cross_f",
            "task": "task_ods_cross_f",
            "dep_type": "宽依赖",
            "project": "CROSS_PROJECT",  # 上游在别的项目
            "group": "CROSS_GROUP",
            "app": "CROSS_APP",
        }]
        out = tmp_path / "schedule_tasks.xlsx"
        generate_schedule_excel(ts_cross, sample_config, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["jobs"]
        # 找依赖行（job类型=tskdep 的那行）
        header = [c.value for c in ws[1]]
        proj_idx = header.index("项目名称")
        group_idx = header.index("任务组名称")
        jobtype_idx = header.index("job类型")
        dep_rows = [row for row in ws.iter_rows(min_row=2, values_only=True)
                    if row[jobtype_idx] == "tskdep"]
        assert dep_rows, "应有依赖行"
        # 依赖行的 project/group 应是上游的 CROSS_PROJECT/CROSS_GROUP
        assert dep_rows[0][proj_idx] == "CROSS_PROJECT", \
            f"跨项目依赖应用上游project，实际={dep_rows[0][proj_idx]}"
        assert dep_rows[0][group_idx] == "CROSS_GROUP", \
            f"跨项目依赖应用上游group，实际={dep_rows[0][group_idx]}"

    def test_same_project_upstream_fallback_current(self, sample_ts, sample_config, tmp_path):
        """同项目上游（upstream 没配 project/group）→ 用当前表的（兜底）。"""
        # sample_ts 的 upstream 没有 project/group → 应兜底用当前表的 SRP_DAILY/GROUP_SPRD
        out = tmp_path / "schedule_tasks.xlsx"
        generate_schedule_excel(sample_ts, sample_config, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["jobs"]
        header = [c.value for c in ws[1]]
        proj_idx = header.index("项目名称")
        jobtype_idx = header.index("job类型")
        dep_rows = [row for row in ws.iter_rows(min_row=2, values_only=True)
                    if row[jobtype_idx] == "tskdep"]
        assert dep_rows
        # 兜底用当前表的配置
        assert dep_rows[0][proj_idx] == "SRP_DAILY", \
            f"同项目应兜底用当前表project，实际={dep_rows[0][proj_idx]}"

    def test_appid_injected_from_config(self, sample_ts, sample_config, tmp_path):
        """appid 经 resolve_config_by_schema 注入 shujia 段，exporter 直接用（不再自查 schema_apps）。"""
        cfg = {
            "shujia": {"appid": "MY_APP_123", "business_owner": "zhangsan"},
            "lts": {"project_name": "SRP_DAILY", "task_group": "GROUP_SPRD"},
        }
        out = tmp_path / "schedule_tasks.xlsx"
        generate_schedule_excel(sample_ts, cfg, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["jobs"]
        header = [c.value for c in ws[1]]
        params_idx = header.index("job参数")
        exec_rows = [row for row in ws.iter_rows(min_row=2, values_only=True)
                     if row[header.index("job类型")] == "url" and row[header.index("job的父节点名称")] == "start"]
        assert exec_rows, "应有执行行"
        assert "MY_APP_123" in str(exec_rows[0][params_idx]), "appid 应注入 job参数"

    def test_taskparams_v_group_code_empty(self, sample_ts, sample_config, tmp_path):
        """V_GROUP_CODE 值留空"""
        out = tmp_path / "schedule_tasks.xlsx"
        generate_schedule_excel(sample_ts, sample_config, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["taskParams"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            param_name = row[3]
            param_value = row[4]
            if param_name == "V_GROUP_CODE":
                assert param_value == "" or param_value is None

    def test_project_group_from_ts_json(self, sample_config, tmp_path):
        """★ 任务四：ts.json 的 task 有 project_name/task_group 时，exporter 直接用（不走 platform_config）。"""
        # 构造 ts，每个 task 带 project_name/task_group（设计阶段确定的）
        ts = {
            "meta": {
                "target": {"f_table": {"schema": "dws", "table": "dwb_test_f"},
                           "i_view": {"schema": "dws", "table": "dwb_test_i"}},
                "schedule": {
                    "cron": "0 30 3 * * ?",
                    "tasks": {
                        "f": {"task_name": "task_dwb_test_f", "job_name": "Pjob_dwb_test_f",
                              "cron": "0 30 3 * * ?", "upstream": [],
                              "project_name": "TS_PROJ_F", "task_group": "TS_GRP_F"},
                        "view": {"task_name": "task_dwb_test_i", "job_name": "Pjob_dwb_test_i",
                                 "cron": "0 30 3 * * ?", "upstream": [],
                                 "project_name": "TS_PROJ_V", "task_group": "TS_GRP_V"},
                        "dq": {"task_name": "task_dwb_test_f_dq", "job_name": "Pjob_dwb_test_f_dq",
                               "cron": "0 30 3 * * ?", "upstream": [],
                               "project_name": "TS_PROJ_DQ", "task_group": "TS_GRP_DQ"},
                    },
                },
            },
            "rules": {},
        }
        # platform_config 的 lts 配的是另一套（验证不被用）
        cfg = {"lts": {"project_name": "PC_PROJ", "task_group": "PC_GRP", "appid": ""},
               "shujia": {}}
        out = tmp_path / "schedule_tasks.xlsx"
        generate_schedule_excel(ts, cfg, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["tasks"]
        header = [c.value for c in ws[1]]
        proj_idx = header.index("项目名称")
        group_idx = header.index("任务组名称")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # F 行用 ts.json 的 TS_PROJ_F
        assert rows[0][proj_idx] == "TS_PROJ_F"
        assert rows[0][group_idx] == "TS_GRP_F"
        # view 行用 TS_PROJ_V
        assert rows[1][proj_idx] == "TS_PROJ_V"
        # dq 行用 TS_PROJ_DQ
        assert rows[2][proj_idx] == "TS_PROJ_DQ"

    def test_project_group_fallback_to_platform_config(self, sample_config, tmp_path):
        """★ 任务四兼容：旧 ts.json 没有 project/task_group -> fallback 到 platform_config 的 lts。"""
        ts = {
            "meta": {
                "target": {"f_table": {"schema": "dws", "table": "dwb_test_f"},
                           "i_view": {"schema": "dws", "table": "dwb_test_i"}},
                "schedule": {
                    "cron": "0 30 3 * * ?",
                    "tasks": {
                        "f": {"task_name": "task_dwb_test_f", "job_name": "Pjob_dwb_test_f",
                              "cron": "0 30 3 * * ?", "upstream": []},  # 无 project/task_group
                    },
                },
            },
            "rules": {},
        }
        out = tmp_path / "schedule_tasks.xlsx"
        generate_schedule_excel(ts, sample_config, out)  # sample_config lts=SRP_DAILY/GROUP_SPRD
        wb = openpyxl.load_workbook(out)
        ws = wb["tasks"]
        header = [c.value for c in ws[1]]
        proj_idx = header.index("项目名称")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert rows[0][proj_idx] == "SRP_DAILY", "旧 ts.json 应回退到 platform_config"


# ============================================================
# 工具函数
# ============================================================

class TestSplitSchemaTable:

    def test_with_schema(self):
        assert _split_schema_table("dws.dwb_xxx_f") == ("dws", "dwb_xxx_f")

    def test_without_schema(self):
        assert _split_schema_table("dwb_xxx_f") == ("", "dwb_xxx_f")
