"""artifact_patcher —— 制品 patch 引擎（docs/specs/opt/07：编辑原语 × 变更声明，严格 patch）。

交付副本模式：读原始制品（provenance 定位）→ 产出**更新后的完整文件副本**进 patched/ +
patch 说明。不就地改原件、不碰仓（推生产不自主）。存量声明漂移不碰（严格 patch）。

add_field 的编辑原语组合（第一刀）：
- 术加 xlsx：TargetFields **行追加**（每落位规则每声明字段一行）+ RULE 表 SQL **单元格替换**
- 代码仓规则组 yml：TargetFields 列表**追加条目** + query_sql 值**替换**（pyyaml round-trip，
  sort_keys=False 保序；已知限制：注释会丢——平台生成 yml 无注释，见 patch 说明）

定位一律稳定标识（rule_code / 目标字段名），不做行号假设；未知列/未知 sheet 不动。
"""
import argparse
import json
import shutil
import sys
from copy import copy
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from run_ut import read_select
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import yaml

# 术加 xlsx 约定列名（与 analyzer/assemble_export 同源的平台约定）
RULE_SHEET_CANDIDATES = ("RULE", "Rule", "rule")
TF_SHEET_CANDIDATES = ("TargetFields", "target_fields", "targetfields")
COL_RULE_CODE = "规则编码"
COL_SQL = "(生成的)查询语句"
TF_COLS = {"规则编码": "rule_code", "目标字段": "target_field", "来源字段": "source_field",
           "加密方式": "encryption", "别名": "alias", "字段类型": "field_type", "备注": "remark"}


def _change_index(ts_v2: dict) -> List[Tuple[str, str, dict]]:
    """[(rule_code, field, 该规则下该字段的定义)] —— TargetFields 行的展开源。"""
    out = []
    for f in (ts_v2.get("change") or {}).get("fields", []):
        for r in f.get("placed_rules", []):
            t = f.get("target_table", "")
            fd = next((x for x in ts_v2.get("tables", {}).get(t, {}).get("fields", [])
                       if x.get("target_field") == f["field"]), {})
            src = f.get("source") or {}
            out.append((r, f["field"], {
                "target_field": f["field"],
                "source_field": f"{src.get('alias', '')}.{src.get('field', f['field'])}".lstrip("."),
                "encryption": "0", "alias": "",
                "field_type": fd.get("field_type", ""),
                "remark": fd.get("field_comment", ""),
            }))
    return out


def _sql_rules(ts_v2: dict) -> List[str]:
    """SQL 单元格需要替换的规则（= 有声明落位的规则）。"""
    return sorted({r for r, _, _ in _change_index(ts_v2)})


# ---------------------------------------------------------------------------
# xlsx patch
# ---------------------------------------------------------------------------

def patch_xlsx(src: Path, dst: Path, ts_v2: dict, etl_dir: Path, notes: List[str]) -> None:
    wb = load_workbook(src)
    rule_ws = next((wb[s] for s in wb.sheetnames if s in RULE_SHEET_CANDIDATES), None)
    tf_ws = next((wb[s] for s in wb.sheetnames if s in TF_SHEET_CANDIDATES), None)
    if rule_ws is None or tf_ws is None:
        raise ValueError(f"xlsx 缺 RULE/TargetFields sheet（现有: {wb.sheetnames}）——"
                         f"provenance 指到的可能不是术加制品包")

    def header_map(ws) -> Dict[str, int]:
        return {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}

    # --- RULE 表：SQL 单元格替换（按 rule_code 定位行） ---
    rh = header_map(rule_ws)
    for rc in _sql_rules(ts_v2):
        sql_text = read_select(etl_dir, rc)   # {code}.sql 或 {code}_描述_模式.sql
        if not sql_text:
            notes.append(f"[跳过] {rc} 无新 SQL 文件，RULE 表未替换")
            continue
        hit = False
        for row in rule_ws.iter_rows(min_row=2):
            if str(row[rh[COL_RULE_CODE] - 1].value).strip() == rc:
                row[rh[COL_SQL] - 1].value = sql_text
                hit = True
                notes.append(f"[替换单元格] RULE.{COL_SQL} @ rule={rc}")
                break
        if not hit:
            notes.append(f"[缺失] RULE 表找不到 rule_code={rc}——定位失败，人工核查")

    # --- TargetFields：行追加（样式从末行复制） ---
    th = header_map(tf_ws)
    for rc, field, vals in _change_index(ts_v2):
        # 已存在（声明漂移/重复执行）→ 不动不覆盖（严格 patch：只落变更清单，不修存量）
        dup = any(str(r[th[COL_RULE_CODE] - 1].value).strip() == rc and
                  str(r[th["目标字段"] - 1].value).strip() == field
                  for r in tf_ws.iter_rows(min_row=2) if r[th[COL_RULE_CODE] - 1].value)
        if dup:
            notes.append(f"[跳过] TargetFields 已有 ({rc}, {field})——不覆盖（严格 patch）")
            continue
        new_row = tf_ws.max_row + 1
        src_row = tf_ws.max_row
        for cn, val in ((COL_RULE_CODE, rc), ("目标字段", vals["target_field"]),
                        ("来源字段", vals["source_field"]), ("加密方式", vals["encryption"]),
                        ("别名", vals["alias"]), ("字段类型", vals["field_type"]),
                        ("备注", vals["remark"])):
            cell = tf_ws.cell(row=new_row, column=th[cn])
            cell.value = val
            src_cell = tf_ws.cell(row=src_row, column=th[cn])
            cell.font = copy(src_cell.font)
            cell.border = copy(src_cell.border)
            cell.alignment = copy(src_cell.alignment)
        notes.append(f"[追加行] TargetFields ({rc}, {field})")
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)


# ---------------------------------------------------------------------------
# 代码仓规则组 yml patch（round-trip 保序；注释丢失为已知限制）
# ---------------------------------------------------------------------------

def patch_yml_group(src_dir: Path, dst_dir: Path, ts_v2: dict, etl_dir: Path,
                    notes: List[str]) -> None:
    dst_dir.mkdir(parents=True, exist_ok=True)
    touched_rules = set(_sql_rules(ts_v2))
    for yml in sorted(src_dir.glob("*.yml")):
        data = yaml.safe_load(yml.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            continue
        rc = str(data.get(COL_RULE_CODE, yml.stem)).strip()
        target_path = dst_dir / yml.name
        if rc in touched_rules:
            sql_text = read_select(etl_dir, rc)
            if sql_text:
                data[COL_SQL] = sql_text
                notes.append(f"[替换单元格] {yml.name}#{COL_SQL}")
            extra = data.get("额外信息（其他sheet页信息）")
            if isinstance(extra, dict) and isinstance(extra.get("TargetFields"), list):
                for _, field, vals in [x for x in _change_index(ts_v2) if x[0] == rc]:
                    dup = any(str(e.get("目标字段", "")).strip() == field
                              for e in extra["TargetFields"]
                              if isinstance(e, dict))
                    if dup:
                        notes.append(f"[跳过] {yml.name} TargetFields 已有 {field}（严格 patch）")
                        continue
                    extra["TargetFields"].append({
                        "规则编码": rc, "目标字段": vals["target_field"],
                        "来源字段": vals["source_field"], "加密方式": vals["encryption"],
                        "别名": vals["alias"], "字段类型": vals["field_type"],
                        "备注": vals["remark"]})
                    notes.append(f"[追加行] {yml.name} TargetFields {field}")
        target_path.write_text(
            yaml.safe_dump(data, allow_unicode=True, sort_keys=False, default_flow_style=False),
            encoding="utf-8")
    notes.append("[说明] yml 为 round-trip 重写（键序保留）；若原文件含注释会丢失——"
                 "平台生成 yml 无注释，真实代码仓如有注释请人工复核")


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="制品 patch：交付副本 + patch 说明（严格 patch）")
    ap.add_argument("--ts-v2", required=True)
    ap.add_argument("--etl-dir", required=True, help="新 SELECT 目录")
    ap.add_argument("--source", required=True,
                    help="原始制品：术加 xlsx 文件 或 代码仓规则组目录（provenance 定位）")
    ap.add_argument("--outdir", required=True, help="patched/ 输出目录")
    args = ap.parse_args(argv)

    ts_v2 = json.loads(Path(args.ts_v2).read_text(encoding="utf-8"))
    if not (ts_v2.get("change") or {}).get("fields"):
        print("PATCH_ERROR: ts 无 change 段", file=sys.stderr)
        return 2
    src = Path(args.source)
    etl_dir = Path(args.etl_dir)
    out = Path(args.outdir)
    notes: List[str] = []
    try:
        if src.is_dir():
            patch_yml_group(src, out / "patched", ts_v2, etl_dir, notes)
        elif src.suffix.lower() in (".xlsx", ".xls"):
            patch_xlsx(src, out / "patched" / src.name, ts_v2, etl_dir, notes)
        else:
            print(f"PATCH_ERROR: 不认识的制品形态 {src}", file=sys.stderr)
            return 2
    except Exception as e:
        print(f"PATCH_ERROR: {e}", file=sys.stderr)
        return 2
    (out / "patch_notes.md").write_text(
        "# patch 说明（严格 patch：只落变更清单）\n\n" + "\n".join(f"- {n}" for n in notes)
        + "\n", encoding="utf-8")
    print(f"patched: {out / 'patched'}")
    print(f"patch_notes: {out / 'patch_notes.md'}（{len(notes)} 条）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
