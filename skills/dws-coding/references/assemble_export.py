#!/usr/bin/env python3
"""
平台制品包 exporter

UT 通过后调用。把验证过的 ts.json + ETL SQL + 视图 DDL 翻译成
内网平台消费的 Excel 格式（execution_tasks.xlsx + schedule_tasks.xlsx）。

产出目录：{outdir}/export/
  - execution_tasks.xlsx   执行平台导入（10 sheet）
  - schedule_tasks.xlsx    调度平台导入（3 sheet）
  - export_manifest.json   元数据清单（给内网 skill 读）

规则编码策略：全部留空，内网部署 skill 回填。
  新建场景：内网先获取编码 → 回填 Excel 三处 → 导入
  优化场景（将来）：编码已有，直接从平台拿

用法:
  python assemble_export.py --ts ts.json --etl-dir etl/ --ddl-dir ddl/ --outdir .

退出码: 0=成功, 1=参数/数据错误, 2=依赖缺失
"""

import sys
import os
import json
import argparse
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("错误: 需要 openpyxl。请运行 pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# ============================================================
# 列定义（精确对齐执行平台制品模板，跟 legacy 一致）
# ============================================================

RULE_COLUMNS = [
    "租户ID", "组织英文简称", "类型", "项目编码", "项目中文名", "项目英文名",
    "项目描述", "子项目编码", "子项目中文名", "子项目英文名", "子项目描述",
    "规则组编码", "规则组中文名称", "规则组英文名称", "规则组业务责任人",
    "规则组描述", "规则组数据源", "规则编码", "规则中文名称", "规则英文名称",
    "创建方式", "规则类型", "数据源", "备注",
    "(生成的）查询语句1", "(生成的）查询语句2", "(生成的）查询语句3",
    "(生成的）查询语句4", "(生成的）查询语句5", "(生成的）查询语句6",
    "(生成的）查询语句7", "(生成的）查询语句8", "(生成的）查询语句9",
    "运行条件", "Select Hint语句", "执行序列", "源Schema", "目标Schema",
    "目标SCHEMA解析值", "目标表", "目标表解析", "是否去重", "删除模式",
    "删除条件", "业务责任人", "delete hint", "交换分区来源表",
    "目标表统计信息收集", "行迁移开关", "会话变量", "环境变量设置",
    "并行开关", "事前操作", "事后操作", "存储模式", "压缩比", "是否散列",
    "程序包名", "SP名称", "API参数", "更新索引", "循环变量",
    "规则循环并行调度标志", "循环分组设置", "循环优先级", "引用规则",
    "重试间隔", "重试次数", "不满足时", "数据库类型", "调度类型",
    "指定分区", "来源表统计分析收集", "统计分析来源表", "规则描述",
    "装载字段", "进程数", "运行内存", "线程数", "批量大小", "并发数",
    "spark数据源",
]
_RULE_COL = {name: idx for idx, name in enumerate(RULE_COLUMNS)}

GROUPVARS_COLUMNS = [
    "规则编码", "动态参数/变量名", "字段类型", "字段定义类型",
    "字段值类型", "变量默认值", "是否校验通过", "数据类型", "描述",
    "是否必填",
]

TARGETFIELDS_COLUMNS = [
    "规则编码", "目标字段名称", "来源字段名称", "加密方式",
    "Merge模式数据源字段值", "别名", "字段类型", "备注",
]

MODELRELATIONS_COLUMNS = [
    "规则编码", "左表schema", "左表名", "左表别名", "右表schema",
    "右表", "右表别名", "模型顺序号", "关联关系", "左表字段列表串",
    "右表字段列表串",
]
EXTRAFIELDS_COLUMNS = ["规则编码", "拓展字段名", "别名", "表达式", "字段类型", "生效", "统计标识"]
SPPARAMS_COLUMNS = ["规则编码", "规则参数名", "数据类型", "入参、出参", "变量默认值"]
CONDITIONS_COLUMNS = [
    "规则编码", "字段名称", "字段关系", "字段值1", "字段值2",
    "与下个条件的逻辑关系", "序号", "字段类型", "条件类型",
    "树形组件业务父类id", "树形组件业务id",
]
MAINTENANCEPARAMS_COLUMNS = ["规则编码", "执行序列", "类型", "schema", "表名", "字段名", "分区表名"]
EXTRACT_COLUMNS = [
    "标签id", "规则编码", "数据库名称", "数据库类型", "标签名",
    "分区读写字段", "分区读写字段类型", "分区数量", "分区下界",
    "分区上界", "批量提取大小", "数据提取SQL", "运行SQL",
    "统计信息分析标识", "统计信息分析来源表信息",
]
EXTRACTCOLUMN_COLUMNS = ["数据标签id", "规则编码", "解密字段", "字段类型", "解密类型"]

# 调度平台列
TASKS_COLUMNS = [
    "项目名称", "任务组名称", "任务名称", "任务类型", "开始时间", "结束时间",
    "调度周期", "依赖上一周期", "日历数据", "责任人", "同步标识", "CTM任务标识",
    "CTM集群标识", "任务是否跳过清场", "TASK资源设置", "不调度过期周期",
    "任务扩展属性", "是否一天多调", "是否并行", "异常任务是否清场",
    "是否导入全量job", "调度频率配置",
]
JOBS_COLUMNS = [
    "项目名称", "任务组名称", "任务名称", "job名称", "job类型",
    "job的父节点名称", "执行路径信息", "job参数", "job调用方法",
    "job超时时间", "job重试次数", "job重试间隔", "job描述",
    "job执行节点", "job变量设置", "job异常处理方式", "job中断处理",
    "job超时处理", "集群名称", "job是否跳过清场", "datastage日志级别",
    "job资源设置", "job扩展属性", "参数空值校验", "生产执行路径",
    "生产job中断处理", "生产schema", "组件资源",
]
TASKPARAMS_COLUMNS = ["项目名称", "任务组名称", "任务名称", "参数名称", "参数值"]

# 审计字段（TargetFields 里过滤掉）
AUDIT_FIELDS = {"del_flag", "crt_cycle_id", "last_upd_cycle_id", "dw_last_update_date"}

# 固定常量
DEFAULT_JOB_PARAMS = '{"headers":"","invokingMode":"同步","returnVal":"","jobRunParams":"","appToken":"","appid":"","authenticationType":"无","timeout":"10"}'
FIXED_PARAMS = ["V_CYCLE_ID", "V_GROUP_CODE"]


# ============================================================
# 配置加载
# ============================================================

def load_platform_config(config_path: str = "") -> dict:
    """读 platform_config.json 原始内容。未找到返回空 dict。

    结构：{ default: {shujia, lts}, schema_mappings: {schema: {shujia, lts}} }
    """
    if not config_path:
        config_path = os.environ.get(
            "PLATFORM_CONFIG",
            str(Path.home() / ".config" / "opencode" / "platform_config.json"),
        )
    p = Path(config_path)
    if not p.exists():
        return {}
    raw = json.loads(p.read_text(encoding="utf-8"))
    # 过滤掉 _comment / _structure 等说明字段
    return {k: v for k, v in raw.items() if not k.startswith("_")}


def resolve_config_by_schema(raw_config: dict, schema: str) -> dict:
    """按 schema 从 platform_config 取两套平台配置。

    查找顺序：schema_mappings[schema] → default
    返回 {shujia: {...}, lts: {...}}
    """
    if not raw_config:
        return {"shujia": {}, "lts": {}}
    default_cfg = raw_config.get("default", {})
    mappings = raw_config.get("schema_mappings", {})
    schema_cfg = mappings.get(schema, {})
    # schema 没配的字段用 default 兜底
    shujia = {**default_cfg.get("shujia", {}), **schema_cfg.get("shujia", {})}
    lts = {**default_cfg.get("lts", {}), **schema_cfg.get("lts", {})}
    return {"shujia": shujia, "lts": lts}


def _cfg(config: dict, key: str, fallback: str = "待配置") -> str:
    """安全取配置值，缺失用 fallback。"""
    val = config.get(key, "")
    return val if val else fallback


# ============================================================
# execution_tasks.xlsx 构建
# ============================================================

def _split_schema_table(full: str) -> tuple[str, str]:
    """schema.table → (schema, table)。无 schema 时返回 ("", full)。"""
    if "." in full:
        sch, tbl = full.rsplit(".", 1)
        return sch, tbl
    return "", full


def build_rule_rows(ts: dict, config: dict, etl_dir: Path, ddl_dir: Path) -> list[list]:
    """构建 RULE sheet 行。顺序：取数规则 → 视图规则 → 参数变量规则。

    config: resolve_config_by_schema 返回的 {shujia, lts} 结构。
    术加执行平台配置从 config["shujia"] 取。
    编码（规则组编码/规则编码）全部留空，内网回填。
    子项目编码留空（schema 对不齐，人工填）。
    """
    rules = ts.get("rules", {})
    meta = ts.get("meta", {})
    f_table = meta.get("target", {}).get("f_table", {})
    i_view = meta.get("target", {}).get("i_view", {})

    target_short = f_table.get("table", "")
    target_full = f"{f_table.get('schema', '')}.{target_short}" if f_table.get("schema") else target_short
    group_desc = f_table.get("cn", "") or target_short

    shujia = config.get("shujia", {})
    data_source = _cfg(shujia, "datasource")
    business_owner = _cfg(shujia, "business_owner", "")
    project_code = _cfg(shujia, "project_code")
    project_cn = _cfg(shujia, "project_cn", project_code)
    project_en = _cfg(shujia, "project_en", project_code)
    # 子项目编码留空（schema 对不齐，人工填）
    sub_code = ""
    sub_cn = ""
    sub_en = ""

    rows = []

    # 公共列填充（每行都要填的项目）
    def _fill_common(row):
        row[_RULE_COL["类型"]] = "3"
        row[_RULE_COL["项目编码"]] = project_code
        row[_RULE_COL["项目中文名"]] = project_cn
        row[_RULE_COL["项目英文名"]] = project_en
        row[_RULE_COL["子项目编码"]] = sub_code
        row[_RULE_COL["子项目中文名"]] = sub_cn
        row[_RULE_COL["子项目英文名"]] = sub_en
        # 规则组编码留空（内网回填）
        row[_RULE_COL["规则组中文名称"]] = target_short
        row[_RULE_COL["规则组英文名称"]] = target_short
        row[_RULE_COL["规则组描述"]] = group_desc
        row[_RULE_COL["规则组数据源"]] = data_source
        row[_RULE_COL["规则组业务责任人"]] = business_owner
        # 规则编码留空（内网回填）

    # --- 取数规则（每条 ETL SQL 一行）---
    for code, rule in rules.items():
        if rule.get("is_view_step"):
            continue
        # 读 ETL SQL 文件
        sql_file = etl_dir / f"{code}.sql"
        if not sql_file.exists():
            # 尝试模糊匹配
            candidates = list(etl_dir.glob(f"*{code}*.sql"))
            sql_file = candidates[0] if candidates else None
        query_sql = sql_file.read_text(encoding="utf-8").strip() if sql_file and sql_file.exists() else ""

        target = rule.get("target_table", "")
        sch, tbl = _split_schema_table(target)

        row = [""] * len(RULE_COLUMNS)
        _fill_common(row)
        row[_RULE_COL["规则中文名称"]] = tbl or target
        row[_RULE_COL["规则英文名称"]] = tbl or target
        row[_RULE_COL["创建方式"]] = "2"
        row[_RULE_COL["规则类型"]] = "1"
        row[_RULE_COL["数据源"]] = data_source
        row[_RULE_COL["备注"]] = "简要描述"
        row[_RULE_COL["(生成的）查询语句1"]] = query_sql
        row[_RULE_COL["运行条件"]] = "0"
        row[_RULE_COL["目标Schema"]] = sch
        row[_RULE_COL["目标表"]] = tbl
        row[_RULE_COL["删除模式"]] = "1"
        row[_RULE_COL["业务责任人"]] = business_owner
        row[_RULE_COL["行迁移开关"]] = "1"
        row[_RULE_COL["并行开关"]] = "0"
        row[_RULE_COL["数据库类型"]] = "GaussDB"
        row[_RULE_COL["调度类型"]] = "0"
        row[_RULE_COL["来源表统计分析收集"]] = "0"
        rows.append(row)

    # --- 视图规则（每个视图 DDL 一行）---
    if i_view and i_view.get("table"):
        # 读视图 DDL
        view_files = list(ddl_dir.glob("create_view_*.sql")) if ddl_dir.exists() else []
        view_ddl = ""
        if view_files:
            view_ddl = view_files[0].read_text(encoding="utf-8").strip()

        view_full = f"{i_view.get('schema', '')}.{i_view.get('table', '')}"
        sch, tbl = _split_schema_table(view_full)

        row = [""] * len(RULE_COLUMNS)
        _fill_common(row)
        row[_RULE_COL["规则中文名称"]] = i_view.get("table", "")
        row[_RULE_COL["规则英文名称"]] = i_view.get("table", "")
        row[_RULE_COL["创建方式"]] = "2"
        row[_RULE_COL["规则类型"]] = "1"
        row[_RULE_COL["数据源"]] = data_source
        row[_RULE_COL["备注"]] = "消费视图封装"
        row[_RULE_COL["(生成的）查询语句1"]] = view_ddl
        row[_RULE_COL["运行条件"]] = "0"
        row[_RULE_COL["目标Schema"]] = sch
        row[_RULE_COL["目标表"]] = tbl
        row[_RULE_COL["删除模式"]] = "0"
        row[_RULE_COL["业务责任人"]] = business_owner
        row[_RULE_COL["行迁移开关"]] = "0"
        row[_RULE_COL["并行开关"]] = "0"
        row[_RULE_COL["数据库类型"]] = "GaussDB"
        row[_RULE_COL["调度类型"]] = "0"
        row[_RULE_COL["来源表统计分析收集"]] = "0"
        rows.append(row)

    # --- 参数变量规则（固定 1 行）---
    pv_row = [""] * len(RULE_COLUMNS)
    _fill_common(pv_row)
    pv_row[_RULE_COL["规则中文名称"]] = "参数变量规则"
    pv_row[_RULE_COL["规则英文名称"]] = "Parameter Variable Rule"
    pv_row[_RULE_COL["创建方式"]] = "1"
    pv_row[_RULE_COL["规则类型"]] = "12"
    pv_row[_RULE_COL["运行条件"]] = "-1"
    pv_row[_RULE_COL["业务责任人"]] = business_owner
    pv_row[_RULE_COL["行迁移开关"]] = "0"
    pv_row[_RULE_COL["并行开关"]] = "0"
    pv_row[_RULE_COL["数据库类型"]] = "GaussDB"
    pv_row[_RULE_COL["调度类型"]] = "0"
    rows.append(pv_row)

    return rows


def build_group_variables(ts: dict) -> list[list]:
    """构建 GroupVariables sheet 行。规则编码留空。

    参数来源：ts.json meta.schedule.exec_params
    """
    exec_params = ts.get("meta", {}).get("schedule", {}).get("exec_params", {})
    rows = []
    for pname in sorted(exec_params.keys()):
        default_val = "19000101000000" if pname == "P_CYCLE_ID" else ""
        rows.append([
            "",           # 规则编码（留空）
            pname,        # 动态参数/变量名
            "1",          # 字段类型
            "1",          # 字段定义类型
            "1",          # 字段值类型
            default_val,  # 变量默认值
            "1",          # 是否校验通过
            "",           # 数据类型
            "",           # 描述
            "",           # 是否必填
        ])
    return rows


def build_target_fields(ts: dict) -> list[list]:
    """构建 TargetFields sheet 行。从 tables 段取字段定义，过滤审计字段。

    字段来源优先级：tables[target_table].fields → rule.fields（旧格式兼容）
    规则编码留空。
    """
    rules = ts.get("rules", {})
    tables = ts.get("tables", {})
    rows = []
    for code, rule in rules.items():
        if rule.get("is_view_step"):
            continue
        # 字段来源：优先 tables 段
        target_tbl = rule.get("target_table", "")
        target_short = target_tbl.rsplit(".", 1)[-1] if "." in target_tbl else target_tbl
        tbl_fields = tables.get(target_short, {}).get("fields", [])
        fields = tbl_fields if tbl_fields else rule.get("fields", [])

        for field in fields:
            target_field = field.get("target_field", "")
            if not target_field or target_field.lower() in AUDIT_FIELDS:
                continue
            # 来源字段：取 source_fields 第一个
            source_fields = field.get("source_fields", [])
            src_field = ""
            src_alias = ""
            if source_fields:
                sf = source_fields[0]
                src_field = sf.get("field", "")
                src_alias = sf.get("alias", "")
            rows.append([
                "",                 # 规则编码（留空）
                target_field,       # 目标字段名称
                src_field,          # 来源字段名称
                "0",                # 加密方式
                "",                 # Merge模式数据源字段值
                src_alias,          # 别名
                "",                 # 字段类型
                "",                 # 备注
            ])
    return rows


def generate_execution_excel(ts: dict, config: dict, etl_dir: Path, ddl_dir: Path, output_path: Path):
    """生成 execution_tasks.xlsx（10 sheet）。"""
    wb = openpyxl.Workbook()

    # Sheet 1: RULE
    ws = wb.active
    ws.title = "RULE"
    ws.append(RULE_COLUMNS)
    for row in build_rule_rows(ts, config, etl_dir, ddl_dir):
        ws.append(row)

    # Sheet 2: GroupVariables
    ws = wb.create_sheet("GroupVariables")
    ws.append(GROUPVARS_COLUMNS)
    for row in build_group_variables(ts):
        ws.append(row)

    # Sheet 3: TargetFields
    ws = wb.create_sheet("TargetFields")
    ws.append(TARGETFIELDS_COLUMNS)
    for row in build_target_fields(ts):
        ws.append(row)

    # Sheet 4-10: 空 sheet（保留表头，跟 legacy 一致）
    empty_sheets = [
        ("ModelRelations", MODELRELATIONS_COLUMNS),
        ("ExtraFields", EXTRAFIELDS_COLUMNS),
        ("SPParams", SPPARAMS_COLUMNS),
        ("Conditions", CONDITIONS_COLUMNS),
        ("MaintenanceParams", MAINTENANCEPARAMS_COLUMNS),
        ("Extract", EXTRACT_COLUMNS),
        ("ExtractColumn", EXTRACTCOLUMN_COLUMNS),
    ]
    for sheet_name, columns in empty_sheets:
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ============================================================
# schedule_tasks.xlsx 构建
# ============================================================

def generate_schedule_excel(ts: dict, config: dict, output_path: Path):
    """生成 schedule_tasks.xlsx（3 sheet）。

    config: resolve_config_by_schema 返回的 {shujia, lts} 结构。
    LTS 调度平台配置从 config["lts"] 取。
    """
    meta = ts.get("meta", {})
    f_table = meta.get("target", {}).get("f_table", {})
    i_view = meta.get("target", {}).get("i_view", {})
    sched = meta.get("schedule", {})

    target_short = f_table.get("table", "")
    lts = config.get("lts", {})
    project_name = _cfg(lts, "project_name")
    task_group = _cfg(lts, "task_group")
    cron = sched.get("cron", "")
    owner = sched.get("owner", "") or _cfg(config.get("shujia", {}), "business_owner", "")
    upstream = sched.get("upstream", [])

    wb = openpyxl.Workbook()

    # --- Sheet 1: tasks ---
    ws = wb.active
    ws.title = "tasks"
    ws.append(TASKS_COLUMNS)

    # F 表任务行
    f_task_name = f"task_{target_short}"
    ws.append([
        project_name, task_group, f_task_name, "周期任务",
        "", "", cron, "是", "", owner,
        "", "", "", "", "", "", "", "", "", "", "", "",
    ])

    # 视图任务行（如有）
    view_task_name = ""
    if i_view and i_view.get("table"):
        view_short = i_view.get("table", "")
        view_task_name = f"task_{view_short}"
        ws.append([
            project_name, task_group, view_task_name, "周期任务",
            "", "", cron, "是", "", owner,
            "", "", "", "", "", "", "", "", "", "", "", "",
        ])

    # --- Sheet 2: jobs ---
    ws = wb.create_sheet("jobs")
    ws.append(JOBS_COLUMNS)

    # F 表执行行
    f_job_name = f"Pjob_{target_short}"
    ws.append([
        project_name, task_group, f_task_name, f_job_name, "url",
        "start", "${V_URL}", DEFAULT_JOB_PARAMS, "",
        "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
    ])

    # 依赖行（tskdep）— 每个 upstream 一行
    for u in upstream:
        dep_task = u.get("task", "")
        if not dep_task:
            continue
        ws.append([
            project_name, task_group, f_task_name, dep_task, "tskdep",
            f_job_name, "", "", "",
            "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
        ])

    # 视图执行行 + 视图依赖 F 表
    if view_task_name:
        view_short = i_view.get("table", "")
        view_job_name = f"Pjob_{view_short}"
        ws.append([
            project_name, task_group, view_task_name, view_job_name, "url",
            "start", "${V_URL}", DEFAULT_JOB_PARAMS, "",
            "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
        ])
        # 视图依赖 F 表的 job
        ws.append([
            project_name, task_group, view_task_name, f_job_name, "tskdep",
            view_job_name, "", "", "",
            "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
        ])

    # --- Sheet 3: taskParams ---
    ws = wb.create_sheet("taskParams")
    ws.append(TASKPARAMS_COLUMNS)

    # 参数从 ts.json 的 lts_params 取（LTS 侧变量名），值留空（内网回填）
    lts_params = sched.get("lts_params", [])
    param_names = [p.get("lts_var", "") for p in lts_params] if lts_params else list(FIXED_PARAMS)

    # 每个任务 × 参数
    all_tasks = [f_task_name]
    if view_task_name:
        all_tasks.append(view_task_name)
    for task_name in all_tasks:
        for param in param_names:
            value = ""  # 值留空（内网回填）
            ws.append([project_name, task_group, task_name, param, value])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ============================================================
# export_manifest.json
# ============================================================

def generate_manifest(ts: dict, config: dict, output_path: Path):
    """生成 export_manifest.json（给内网 skill 读的元数据清单）。"""
    meta = ts.get("meta", {})
    f_table = meta.get("target", {}).get("f_table", {})
    i_view = meta.get("target", {}).get("i_view", {})
    sched = meta.get("schedule", {})
    rules = ts.get("rules", {})

    target_short = f_table.get("table", "")
    target_full = f"{f_table.get('schema', '')}.{target_short}" if f_table.get("schema") else target_short
    has_view = bool(i_view and i_view.get("table"))

    # 需要的规则编码数 = 取数规则数 + 视图规则数 + 1(参数变量)
    etl_count = sum(1 for r in rules.values() if not r.get("is_view_step"))
    view_count = 1 if has_view else 0
    rule_codes_needed = etl_count + view_count + 1

    upstream_tasks = []
    for u in sched.get("upstream", []):
        upstream_tasks.append({
            "source_table": u.get("table", ""),
            "schedule_task": u.get("task", ""),
        })

    manifest = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "target_table": target_full,
        "target_table_short": target_short,
        "view_name": f"{i_view.get('schema', '')}.{i_view.get('table', '')}" if has_view else "",
        "task_name": f"task_{target_short}",
        "job_name": f"Pjob_{target_short}",
        "view_task_name": f"task_{i_view.get('table', '')}" if has_view else "",
        "view_job_name": f"Pjob_{i_view.get('table', '')}" if has_view else "",
        "cron_expr": sched.get("cron", ""),
        "project_name": _cfg(config.get("lts", {}), "project_name"),
        "task_group": _cfg(config.get("lts", {}), "task_group"),
        "params": sorted(ts.get("meta", {}).get("schedule", {}).get("exec_params", {}).keys()),
        "upstream_tasks": upstream_tasks,
        "rule_codes_needed": rule_codes_needed,
        "codes_filled": False,   # 内网回填后改 true
        "files": ["execution_tasks.xlsx", "schedule_tasks.xlsx"],
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


# ============================================================
# 主入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="平台制品包 exporter（UT 通过后调用）")
    parser.add_argument("--ts", required=True, help="ts.json 路径")
    parser.add_argument("--etl-dir", required=True, help="ETL SQL 目录（etl/）")
    parser.add_argument("--ddl-dir", required=True, help="DDL 目录（ddl/）")
    parser.add_argument("--outdir", required=True, help="产出根目录（export/ 建在此下）")
    parser.add_argument("--config", default="", help="platform_config.json 路径")
    args = parser.parse_args()

    ts_path = Path(args.ts)
    etl_dir = Path(args.etl_dir)
    ddl_dir = Path(args.ddl_dir)
    export_dir = Path(args.outdir) / "export"

    # 读 ts.json
    if not ts_path.exists():
        print(f"错误: ts.json 不存在: {ts_path}", file=sys.stderr)
        sys.exit(1)
    ts = json.loads(ts_path.read_text(encoding="utf-8"))

    # 读配置（按目标表 schema 映射两套平台配置）
    raw_config = load_platform_config(args.config)
    target_schema = ts.get("meta", {}).get("target", {}).get("f_table", {}).get("schema", "")
    config = resolve_config_by_schema(raw_config, target_schema)

    # 产出
    exec_path = export_dir / "execution_tasks.xlsx"
    sched_path = export_dir / "schedule_tasks.xlsx"
    manifest_path = export_dir / "export_manifest.json"

    generate_execution_excel(ts, config, etl_dir, ddl_dir, exec_path)
    generate_schedule_excel(ts, config, sched_path)
    generate_manifest(ts, config, manifest_path)

    print("=" * 50)
    print("平台制品包已生成:")
    print(f"  {exec_path}")
    print(f"  {sched_path}")
    print(f"  {manifest_path}")
    print(f"  规则编码需求: {manifest_path and json.loads(manifest_path.read_text(encoding='utf-8'))['rule_codes_needed']} 个（留空，内网回填）")
    print(f"  目标表: {ts.get('meta', {}).get('target', {}).get('f_table', {}).get('table', '')}")


if __name__ == "__main__":
    main()
