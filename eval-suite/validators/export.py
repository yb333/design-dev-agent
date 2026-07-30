"""Export validator — validate export Excel structure."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .base import BaseValidator, CheckResult, CheckStatus


class ExportValidator(BaseValidator):
    """Validate export Excel files existence and sheet structure."""

    @property
    def name(self) -> str:
        return "export"

    def validate(
        self,
        check_config: dict,
        output_dir: Path,
        golden_dir: Path | None = None,
    ) -> list[CheckResult]:
        check_type = check_config.get("type", "")

        if check_type == "export_files_exist":
            return self._check_files_exist(output_dir)
        if check_type == "export_sheets":
            return self._check_sheets(check_config, output_dir)

        return [
            CheckResult(
                check_type=check_type,
                status=CheckStatus.SKIP,
                detail=f"Unknown check type: {check_type}",
            )
        ]

    def _check_files_exist(self, output_dir: Path) -> list[CheckResult]:
        export_dir = output_dir / "09_export"
        if not export_dir.exists():
            return [
                CheckResult(
                    check_type="export_files_exist",
                    status=CheckStatus.FAIL,
                    detail="Export directory not found",
                    evidence=str(export_dir),
                )
            ]

        results: list[CheckResult] = []
        expected_files = ["execution_tasks.xlsx", "schedule_tasks.xlsx"]

        for filename in expected_files:
            path = export_dir / filename
            if path.exists() and path.stat().st_size > 0:
                results.append(
                    CheckResult(
                        check_type="export_files_exist",
                        status=CheckStatus.PASS,
                        detail=f"Export file exists: {filename}",
                    )
                )
            else:
                reason = "not found" if not path.exists() else "empty (0 bytes)"
                results.append(
                    CheckResult(
                        check_type="export_files_exist",
                        status=CheckStatus.FAIL,
                        detail=f"Export file {reason}: {filename}",
                        evidence=str(path),
                    )
                )

        return results

    def _check_sheets(
        self, check_config: dict, output_dir: Path
    ) -> list[CheckResult]:
        file_path = check_config.get("file", "")
        expected_sheets = check_config.get("sheets", [])

        if not file_path:
            return [
                CheckResult(
                    check_type="export_sheets",
                    status=CheckStatus.FAIL,
                    detail="No 'file' specified in config",
                )
            ]

        if not expected_sheets:
            return [
                CheckResult(
                    check_type="export_sheets",
                    status=CheckStatus.SKIP,
                    detail="No 'sheets' specified in config",
                )
            ]

        full_path = output_dir / file_path
        if not full_path.exists():
            return [
                CheckResult(
                    check_type="export_sheets",
                    status=CheckStatus.FAIL,
                    detail=f"Excel file not found: {file_path}",
                    evidence=str(full_path),
                )
            ]

        try:
            wb = load_workbook(str(full_path), read_only=True)
            actual_sheets = wb.sheetnames
            wb.close()
        except Exception as e:
            return [
                CheckResult(
                    check_type="export_sheets",
                    status=CheckStatus.FAIL,
                    detail=f"Failed to read Excel file: {file_path}",
                    evidence=str(e),
                )
            ]

        results: list[CheckResult] = []
        for sheet_name in expected_sheets:
            if sheet_name in actual_sheets:
                results.append(
                    CheckResult(
                        check_type="export_sheets",
                        status=CheckStatus.PASS,
                        detail=f"Sheet '{sheet_name}' found in {file_path}",
                    )
                )
            else:
                results.append(
                    CheckResult(
                        check_type="export_sheets",
                        status=CheckStatus.FAIL,
                        detail=(
                            f"Sheet '{sheet_name}' not found in {file_path}. "
                            f"Available: {actual_sheets}"
                        ),
                        evidence=f"Expected: {sheet_name}, Available: {actual_sheets}",
                    )
                )

        return results
