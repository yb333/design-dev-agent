"""Golden diff validator — structural comparison against golden output."""

from __future__ import annotations

from pathlib import Path

from .base import BaseValidator, CheckResult, CheckStatus
from .design import _extract_field_mappings, _extract_segments
from .review import extract_review_stats
from .sql import parse_ddl, parse_etl, _normalize_type

# Mapping of output subdirectories to their comparison handlers
_DIR_HANDLERS = {
    "02_design": "_compare_design",
    "04_ddl": "_compare_ddl",
    "05_etl": "_compare_etl",
    "07_code_review": "_compare_review",
    "09_export": "_compare_export",
}


def _list_common_files(dir_a: Path, dir_b: Path) -> list[str]:
    """List filenames that exist in both directories."""
    if not dir_a.exists() or not dir_b.exists():
        return []
    set_a = {f.name for f in dir_a.iterdir() if f.is_file()}
    set_b = {f.name for f in dir_b.iterdir() if f.is_file()}
    return sorted(set_a & set_b)


class GoldenDiffValidator(BaseValidator):
    """Compare actual output against golden (human-confirmed) output."""

    @property
    def name(self) -> str:
        return "golden_diff"

    def validate(
        self,
        check_config: dict,
        output_dir: Path,
        golden_dir: Path | None = None,
    ) -> list[CheckResult]:
        check_type = check_config.get("type", "")

        if check_type == "golden_structure_match":
            resolved_golden = self._resolve_golden_dir(check_config, output_dir)
            if not resolved_golden or not resolved_golden.exists():
                return [
                    CheckResult(
                        check_type="golden_structure_match",
                        status=CheckStatus.FAIL,
                        detail=f"Golden directory not found: {resolved_golden}",
                    )
                ]
            return self._compare_all(output_dir, resolved_golden)

        return [
            CheckResult(
                check_type=check_type,
                status=CheckStatus.SKIP,
                detail=f"Unknown check type: {check_type}",
            )
        ]

    def _resolve_golden_dir(
        self, check_config: dict, output_dir: Path
    ) -> Path | None:
        """Resolve golden directory path from config.

        The golden_dir in config is relative to eval-suite root.
        """
        golden_rel = check_config.get("golden_dir", "")
        if not golden_rel:
            return None

        # Try resolving relative to eval-suite root (parent of validators/)
        eval_suite_root = Path(__file__).resolve().parent.parent
        resolved = eval_suite_root / golden_rel
        return resolved

    def _compare_all(
        self, output_dir: Path, golden_dir: Path
    ) -> list[CheckResult]:
        """Compare all applicable subdirectories."""
        results: list[CheckResult] = []

        for subdir, handler_name in _DIR_HANDLERS.items():
            actual_sub = output_dir / subdir
            golden_sub = golden_dir / subdir

            if not actual_sub.exists() or not golden_sub.exists():
                continue

            handler = getattr(self, handler_name, None)
            if handler is None:
                continue

            sub_results = handler(actual_sub, golden_sub)
            results.extend(sub_results)

        return results

    def _compare_design(
        self, actual_dir: Path, golden_dir: Path
    ) -> list[CheckResult]:
        """Compare design.md files."""
        actual_design = actual_dir / "design.md"
        golden_design = golden_dir / "design.md"

        if not actual_design.exists() or not golden_design.exists():
            return []

        actual_text = actual_design.read_text(encoding="utf-8")
        golden_text = golden_design.read_text(encoding="utf-8")

        results: list[CheckResult] = []

        # Compare field mappings
        actual_mappings = _extract_field_mappings(actual_text)
        golden_mappings = _extract_field_mappings(golden_text)

        if actual_mappings == golden_mappings:
            results.append(
                CheckResult(
                    check_type="golden_structure_match",
                    status=CheckStatus.PASS,
                    detail=f"design.md field mappings match ({len(actual_mappings)} pairs)",
                    evidence="file: 02_design/design.md",
                )
            )
        else:
            missing = golden_mappings - actual_mappings
            extra = actual_mappings - golden_mappings
            results.append(
                CheckResult(
                    check_type="golden_structure_match",
                    status=CheckStatus.FAIL,
                    detail=(
                        f"design.md field mappings differ: "
                        f"actual={len(actual_mappings)}, golden={len(golden_mappings)}, "
                        f"missing={len(missing)}, extra={len(extra)}"
                    ),
                    evidence=f"file: 02_design/design.md | missing={sorted(missing)} | extra={sorted(extra)}",
                )
            )

        # Compare segment count
        actual_segs = _extract_segments(actual_text)
        golden_segs = _extract_segments(golden_text)

        if actual_segs == golden_segs:
            results.append(
                CheckResult(
                    check_type="golden_structure_match",
                    status=CheckStatus.PASS,
                    detail=f"design.md segment count matches ({actual_segs})",
                    evidence="file: 02_design/design.md",
                )
            )
        else:
            results.append(
                CheckResult(
                    check_type="golden_structure_match",
                    status=CheckStatus.FAIL,
                    detail=(
                        f"design.md segment count differs: "
                        f"actual={actual_segs}, golden={golden_segs}"
                    ),
                    evidence="file: 02_design/design.md",
                )
            )

        return results

    def _compare_ddl(
        self, actual_dir: Path, golden_dir: Path
    ) -> list[CheckResult]:
        """Compare DDL SQL files."""
        common_files = _list_common_files(actual_dir, golden_dir)
        results: list[CheckResult] = []

        for filename in common_files:
            if not filename.endswith(".sql"):
                continue

            actual_sql = (actual_dir / filename).read_text(encoding="utf-8")
            golden_sql = (golden_dir / filename).read_text(encoding="utf-8")

            actual_parsed = parse_ddl(actual_sql)
            golden_parsed = parse_ddl(golden_sql)

            if actual_parsed is None or golden_parsed is None:
                results.append(
                    CheckResult(
                        check_type="golden_structure_match",
                        status=CheckStatus.FAIL,
                        detail=f"DDL parse error: {filename} (actual or golden)",
                        evidence=f"file: 04_ddl/{filename}",
                    )
                )
                continue

            actual_cols = actual_parsed["columns"]
            golden_cols = golden_parsed["columns"]

            # Normalize types for comparison
            actual_norm = {k: _normalize_type(v) for k, v in actual_cols.items()}
            golden_norm = {k: _normalize_type(v) for k, v in golden_cols.items()}

            if actual_norm == golden_norm:
                results.append(
                    CheckResult(
                        check_type="golden_structure_match",
                        status=CheckStatus.PASS,
                        detail=f"DDL columns match: {filename} ({len(actual_norm)} columns)",
                        evidence=f"file: 04_ddl/{filename}",
                    )
                )
            else:
                missing = sorted(set(golden_norm) - set(actual_norm))
                extra = sorted(set(actual_norm) - set(golden_norm))
                # Type mismatches in common columns
                common = set(actual_norm) & set(golden_norm)
                type_diffs = [
                    f"{c}: {actual_norm[c]} != {golden_norm[c]}"
                    for c in sorted(common)
                    if actual_norm[c] != golden_norm[c]
                ]
                detail_parts = []
                if missing:
                    detail_parts.append(f"missing_cols={missing}")
                if extra:
                    detail_parts.append(f"extra_cols={extra}")
                if type_diffs:
                    detail_parts.append(f"type_diffs={type_diffs}")

                results.append(
                    CheckResult(
                        check_type="golden_structure_match",
                        status=CheckStatus.FAIL,
                        detail=f"DDL columns differ: {filename} — {'; '.join(detail_parts)}",
                        evidence=f"file: 04_ddl/{filename}",
                    )
                )

        return results

    def _compare_etl(
        self, actual_dir: Path, golden_dir: Path
    ) -> list[CheckResult]:
        """Compare ETL SQL files."""
        common_files = _list_common_files(actual_dir, golden_dir)
        results: list[CheckResult] = []

        for filename in common_files:
            if not filename.endswith(".sql"):
                continue

            actual_sql = (actual_dir / filename).read_text(encoding="utf-8")
            golden_sql = (golden_dir / filename).read_text(encoding="utf-8")

            actual_parsed = parse_etl(actual_sql)
            golden_parsed = parse_etl(golden_sql)

            if actual_parsed is None or golden_parsed is None:
                results.append(
                    CheckResult(
                        check_type="golden_structure_match",
                        status=CheckStatus.FAIL,
                        detail=f"ETL parse error: {filename} (actual or golden)",
                        evidence=f"file: 05_etl/{filename}",
                    )
                )
                continue

            actual_select = set(actual_parsed.get("select_columns", []))
            golden_select = set(golden_parsed.get("select_columns", []))

            def _bare_col(name: str) -> str:
                return name.split(".")[-1].strip()

            actual_bare = {_bare_col(c) for c in actual_select}
            golden_bare = {_bare_col(c) for c in golden_select}

            if actual_bare == golden_bare:
                results.append(
                    CheckResult(
                        check_type="golden_structure_match",
                        status=CheckStatus.PASS,
                        detail=f"ETL SELECT columns match: {filename} ({len(actual_select)} columns)",
                        evidence=f"file: 05_etl/{filename}",
                    )
                )
            else:
                missing = sorted(golden_bare - actual_bare)
                extra = sorted(actual_bare - golden_bare)
                results.append(
                    CheckResult(
                        check_type="golden_structure_match",
                        status=CheckStatus.FAIL,
                        detail=(
                            f"ETL SELECT columns differ: {filename} — "
                            f"actual={len(actual_bare)}, golden={len(golden_bare)}, "
                            f"missing={len(missing)}, extra={len(extra)}"
                        ),
                        evidence=f"file: 05_etl/{filename} | missing={missing} | extra={extra}",
                    )
                )

        return results

    def _compare_review(
        self, actual_dir: Path, golden_dir: Path
    ) -> list[CheckResult]:
        """Compare review markdown files."""
        common_files = _list_common_files(actual_dir, golden_dir)
        results: list[CheckResult] = []

        for filename in common_files:
            if not filename.endswith(".md"):
                continue

            actual_text = (actual_dir / filename).read_text(encoding="utf-8")
            golden_text = (golden_dir / filename).read_text(encoding="utf-8")

            actual_stats = extract_review_stats(actual_text)
            golden_stats = extract_review_stats(golden_text)

            if actual_stats is None or golden_stats is None:
                results.append(
                    CheckResult(
                        check_type="golden_structure_match",
                        status=CheckStatus.SKIP,
                        detail=f"Could not extract review stats from {filename}",
                        evidence=f"file: 07_code_review/{filename}",
                    )
                )
                continue

            # Compare CRITICAL and MAJOR counts
            counts_match = (
                actual_stats["critical"] == golden_stats["critical"]
                and actual_stats["major"] == golden_stats["major"]
            )

            if counts_match:
                results.append(
                    CheckResult(
                        check_type="golden_structure_match",
                        status=CheckStatus.PASS,
                        detail=(
                            f"Review counts match: {filename} "
                            f"(CRITICAL={actual_stats['critical']}, "
                            f"MAJOR={actual_stats['major']})"
                        ),
                        evidence=f"file: 07_code_review/{filename}",
                    )
                )
            else:
                results.append(
                    CheckResult(
                        check_type="golden_structure_match",
                        status=CheckStatus.FAIL,
                        detail=(
                            f"Review counts differ: {filename} — "
                            f"actual(CRITICAL={actual_stats['critical']}, "
                            f"MAJOR={actual_stats['major']}), "
                            f"golden(CRITICAL={golden_stats['critical']}, "
                            f"MAJOR={golden_stats['major']})"
                        ),
                        evidence=f"file: 07_code_review/{filename}",
                    )
                )

        return results

    def _compare_export(
        self, actual_dir: Path, golden_dir: Path
    ) -> list[CheckResult]:
        """Compare export Excel files (structural — check they both exist)."""
        common_files = _list_common_files(actual_dir, golden_dir)
        results: list[CheckResult] = []

        for filename in common_files:
            if not filename.endswith(".xlsx"):
                continue

            # Structural comparison: check that both files have same sheet names
            from openpyxl import load_workbook

            try:
                actual_wb = load_workbook(str(actual_dir / filename), read_only=True)
                golden_wb = load_workbook(str(golden_dir / filename), read_only=True)
                actual_sheets = set(actual_wb.sheetnames)
                golden_sheets = set(golden_wb.sheetnames)
                actual_wb.close()
                golden_wb.close()
            except Exception as e:
                results.append(
                    CheckResult(
                        check_type="golden_structure_match",
                        status=CheckStatus.FAIL,
                        detail=f"Failed to read Excel: {filename}",
                        evidence=str(e),
                    )
                )
                continue

            if actual_sheets == golden_sheets:
                results.append(
                    CheckResult(
                        check_type="golden_structure_match",
                        status=CheckStatus.PASS,
                        detail=f"Export sheets match: {filename} ({sorted(actual_sheets)})",
                        evidence=f"file: 09_export/{filename}",
                    )
                )
            else:
                missing = sorted(golden_sheets - actual_sheets)
                extra = sorted(actual_sheets - golden_sheets)
                results.append(
                    CheckResult(
                        check_type="golden_structure_match",
                        status=CheckStatus.FAIL,
                        detail=(
                            f"Export sheets differ: {filename} — "
                            f"missing={missing}, extra={extra}"
                        ),
                        evidence=f"file: 09_export/{filename}",
                    )
                )

        return results
