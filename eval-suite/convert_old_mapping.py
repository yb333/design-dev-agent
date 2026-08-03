#!/usr/bin/env python3
"""
旧格式 mapping → 新格式 mapping + RS.md 转换器

表头严格使用标准模板（docs/templates/mapping模板.xlsx），不自己写。
以标准模板为骨架填充数据。

用法:
  python convert_old_mapping.py --input 旧mapping.xlsx --outdir cases/xxx/
"""

import sys
import shutil
import argparse
from pathlib import Path
import openpyxl

TEMPLATE = Path(__file__).resolve().parent.parent / "docs" / "templates" / "mapping模板.xlsx"

# 标准模板的实体级列顺序（按模板实际表头，含*号）
ENTITY_HEADER = [
    "序号", "分组", "源表schema*", "源表中文名", "源表物理表名*",
    "源表别名*", "目标表逻辑schema*", "目标表中文名", "目标表物理名称*",
    "取数规则", "关联&限定条件", "备注", "数据库类型"
]

# 标准模板的属性级列顺序
ATTR_HEADER = [
    "序号", "分组", "源Schema", "源表物理表名", "源表物理表别名",
    "源表字段中文名", "源表字段名", "源表字段类型",
    "映射规则*", "映射表达式", "目标字段名*", "目标字段中文名", "目标字段类型",
    "备注", "数据标准"
]

RULE_MAP = {"直取": "直接复制", "加工": "数据加工", "赋值": "赋值", "序列": "序列"}

STANDARD_AUDIT = [
    ("del_flag", "删除标识", "NVARCHAR(1)", "'N'", "审计字段"),
    ("crt_cycle_id", "创建批次ID", "BIGINT", "'${P_CYCLE_ID}'", "审计字段"),
    ("last_upd_cycle_id", "最后更新批次ID", "BIGINT", "'${P_CYCLE_ID}'", "审计字段"),
    ("dw_last_update_date", "数仓最后更新时间", "TIMESTAMP(0) WITHOUT TIME ZONE", "CURRENT_TIMESTAMP", "审计字段"),
]


def auto_alias(table_name: str, existing: set, index: int) -> str:
    parts = table_name.replace(".", "_").split("_")
    alias = "".join(p[0] for p in parts if p)[:3] if len(parts) > 1 else parts[0][:2]
    base = alias
    while alias in existing:
        alias = f"{base}{index}"
    existing.add(alias)
    return alias


def read_old_entity(ws):
    """读旧格式实体级，返回 list[dict]"""
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        cells = [str(c) if c is not None else "" for c in row]
        if not cells[1]:
            continue
        rows.append({
            "src_schema": cells[0], "src_table": cells[1], "src_cn": cells[2],
            "tgt_schema": cells[3] if len(cells) > 3 else "",
            "tgt_cn": cells[4] if len(cells) > 4 else "",
            "tgt_table": cells[5] if len(cells) > 5 else "",
            "join_cond": cells[6] if len(cells) > 6 else "",
            "remark": cells[7] if len(cells) > 7 else "",
        })
    return rows


def read_old_attr(ws):
    """读旧格式属性级，返回 list[dict]"""
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        cells = [str(c) if c is not None else "" for c in row]
        if len(cells) < 8 or not cells[6]:
            continue
        rows.append({
            "src_schema": cells[0], "src_table": cells[1],
            "src_field": cells[2], "src_type": cells[3],
            "old_rule": cells[4] if len(cells) > 4 else "直取",
            "expression": cells[5] if len(cells) > 5 else "",
            "target_field": cells[6] if len(cells) > 6 else "",
            "target_cn": cells[7] if len(cells) > 7 else "",
            "target_type": cells[8] if len(cells) > 8 else "",
        })
    return rows


def convert(input_path: str, outdir: str):
    # 读旧文件（用 pandas 兜底 openpyxl 格式问题）
    try:
        wb_old = openpyxl.load_workbook(input_path, data_only=True)
    except Exception:
        import pandas as pd
        wb_old = openpyxl.Workbook()
        for sn in ["实体级mapping", "属性级mapping"]:
            df = pd.read_excel(input_path, sheet_name=sn)
            ws = wb_old.create_sheet(sn)
            for col in df.columns:
                ws.append([col])
            for _, row in df.iterrows():
                ws.append([row[col] if pd.notna(row[col]) else "" for col in df.columns])
        if "Sheet" in wb_old.sheetnames:
            del wb_old["Sheet"]

    # 读旧数据
    old_entity = read_old_entity(wb_old["实体级mapping"])
    old_attr = read_old_attr(wb_old["属性级mapping"])

    if not old_entity:
        print("错误：实体级mapping无数据", file=sys.stderr)
        return

    # 生成别名
    aliases = set()
    for i, e in enumerate(old_entity):
        e["alias"] = auto_alias(e["src_table"], aliases, i)
    table_to_alias = {e["src_table"]: e["alias"] for e in old_entity}

    # 以标准模板为骨架，复制一份再填数据
    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    mapping_path = outdir / "mapping.xlsx"
    shutil.copy2(str(TEMPLATE), str(mapping_path))

    wb = openpyxl.load_workbook(str(mapping_path))

    # === 实体级：清空模板样例行，填入转换数据 ===
    ws1 = wb["实体级mapping"]
    # 删除模板的数据行（第2行到最后）
    for row in range(ws1.max_row, 1, -1):
        ws1.delete_rows(row)

    for idx, e in enumerate(old_entity, 1):
        ws1.append([
            idx,                    # 序号
            "default",              # 分组
            e["src_schema"],        # 源表schema*
            e["src_cn"],            # 源表中文名
            e["src_table"],         # 源表物理表名*
            e["alias"],             # 源表别名*
            e["tgt_schema"] or old_entity[0]["tgt_schema"],  # 目标表逻辑schema*
            e["tgt_cn"] or old_entity[0]["tgt_cn"],          # 目标表中文名
            e["tgt_table"] or old_entity[0]["tgt_table"],    # 目标表物理名称*
            "",                     # 取数规则
            e["join_cond"],         # 关联&限定条件
            e["remark"],            # 备注
            "",                     # 数据库类型
        ])

    # === 属性级：清空模板样例行，填入转换数据 ===
    ws2 = wb["属性级mapping"]
    for row in range(ws2.max_row, 1, -1):
        ws2.delete_rows(row)

    # 检查审计字段
    existing_targets = {a["target_field"].lower() for a in old_attr}

    seq = 1
    for a in old_attr:
        rule = RULE_MAP.get(a["old_rule"], "直接复制")
        alias = table_to_alias.get(a["src_table"], "")
        remark = "主键" if seq == 1 else ""
        ws2.append([
            seq,                # 序号
            "default",          # 分组
            a["src_schema"],    # 源Schema
            a["src_table"],     # 源表物理表名
            alias,              # 源表物理表别名
            a["src_field"],     # 源表字段中文名
            a["src_field"],     # 源表字段名
            a["src_type"],      # 源表字段类型
            rule,               # 映射规则*
            a["expression"] if a["expression"] else "-",  # 映射表达式
            a["target_field"],  # 目标字段名*
            a["target_cn"],     # 目标字段中文名
            a["target_type"],   # 目标字段类型
            remark,             # 备注
            "",                 # 数据标准
        ])
        seq += 1

    # 补充审计字段
    if not any(audit[0] in existing_targets for audit in STANDARD_AUDIT):
        for aname, acn, atype, aexpr, aremark in STANDARD_AUDIT:
            ws2.append([seq, "default", "", "", "", "", "", "",
                        "赋值", aexpr, aname, acn, atype, aremark, ""])
            seq += 1

    wb.save(str(mapping_path))

    # 生成 RS.md
    tgt_table = old_entity[0]["tgt_table"]
    tgt_schema = old_entity[0]["tgt_schema"]
    tgt_cn = old_entity[0]["tgt_cn"]
    rs_path = outdir / "RS.md"
    rs_path.write_text(f"""# RS - {tgt_cn}

## 1.1 资产基本信息

| 属性 | 内容 |
|------|------|
| SCHEMA | {tgt_schema} |
| 资产名称 | {tgt_table} |
| 资产描述 | {tgt_cn} |
| 业务对象 | {tgt_cn.replace('宽表', '').replace('中心', '')} |
| 逻辑数据实体 | 每行一个{tgt_cn.replace('宽表', '').replace('中心', '')}记录 |
| owner 部门 | 数据开发部 |
| owner 人员 | zhangsan |

## L07 初始化及调度

| 配置项 | 内容 |
|--------|------|
| 调度方案 | 全量调度 |
| 调度频率 | T+1，一天一调 |
| 调度完成时间 | 3:30 |
| 增量识别 | 不涉及 |
""", encoding="utf-8")

    n_audit = 0 if any(audit[0] in existing_targets for audit in STANDARD_AUDIT) else 4
    print(f"✅ {tgt_table}: {len(old_attr)}业务字段 + {n_audit}审计 = {len(old_attr)+n_audit}总, {len(old_entity)}源表")


def main():
    parser = argparse.ArgumentParser(description="旧格式mapping→新格式+RS转换（表头严格按标准模板）")
    parser.add_argument("--input", required=True, help="旧格式mapping.xlsx路径")
    parser.add_argument("--outdir", required=True, help="输出目录")
    args = parser.parse_args()
    convert(args.input, args.outdir)


if __name__ == "__main__":
    main()
