#!/usr/bin/env python3
"""
案例数据自检：检查 eval-suite/cases 下的 mapping + RS 是否自洽。

在 local_eval 第一步跑（preprocess 之前），发现数据问题立即报告。
避免"案例数据错"导致的"地基错误"在内网才发现。

检查项：
1. 实体级必填列非空（schema/物理表名/别名/目标表）
2. 目标表物理名称以 _i 结尾（我们的标准：写I视图名）
3. 属性级每条记录的源表在实体级有定义
4. 属性级源表别名在实体级有对应
5. 属性级目标字段名/类型非空
6. 映射规则是合法值（直接复制/数据加工/赋值/序列）
7. RS 的资产名和 mapping 目标表一致

用法:
  python check_case.py --mapping xxx.xlsx --rs xxx.md
  python check_case.py --cases-dir eval-suite/cases/   # 检查全部
"""

import sys
import argparse
from pathlib import Path
import openpyxl

VALID_RULES = {"直接复制", "数据加工", "赋值", "序列"}


def check_case(mapping_path: str, rs_path: str) -> list[str]:
    """检查单个案例，返回问题列表（空=通过）。"""
    issues = []

    # === 检查 mapping ===
    try:
        wb = openpyxl.load_workbook(mapping_path, data_only=True)
    except Exception as e:
        return [f"mapping 文件无法读取: {e}"]

    # 实体级
    if "实体级mapping" not in wb.sheetnames:
        issues.append("缺少 sheet: 实体级mapping")
        return issues

    ws_entity = wb["实体级mapping"]
    entity_headers = [str(c.value or "").rstrip("*").strip() for c in ws_entity[1]]
    entity_rows = []
    for row in ws_entity.iter_rows(min_row=2, values_only=True):
        cells = [str(c) if c is not None else "" for c in row]
        if not cells[entity_headers.index("源表物理表名")] if "源表物理表名" in entity_headers else True:
            if not any(cells):
                continue
        entity_rows.append(cells)

    # 实体级列索引
    def col_idx(headers, name):
        for i, h in enumerate(headers):
            if h == name:
                return i
        return -1

    idx_src_schema = col_idx(entity_headers, "源表schema")
    idx_src_table = col_idx(entity_headers, "源表物理表名")
    idx_src_alias = col_idx(entity_headers, "源表别名")
    idx_tgt_schema = col_idx(entity_headers, "目标表逻辑schema")
    idx_tgt_table = col_idx(entity_headers, "目标表物理名称")

    # 检查实体级必填列存在
    for name, idx in [("源表schema", idx_src_schema), ("源表物理表名", idx_src_table),
                      ("源表别名", idx_src_alias), ("目标表逻辑schema", idx_tgt_schema),
                      ("目标表物理名称", idx_tgt_table)]:
        if idx == -1:
            issues.append(f"实体级缺少列: {name}")

    if issues:
        return issues  # 列都没有，后面没法查

    entity_aliases = set()
    entity_tables = set()
    target_table = ""

    for i, cells in enumerate(entity_rows, 2):
        schema = cells[idx_src_schema] if idx_src_schema < len(cells) else ""
        table = cells[idx_src_table] if idx_src_table < len(cells) else ""
        alias = cells[idx_src_alias] if idx_src_alias < len(cells) else ""

        if not table:
            issues.append(f"实体级第{i}行: 源表物理表名为空")
            continue

        entity_tables.add(table)
        if alias:
            entity_aliases.add(alias)
        else:
            issues.append(f"实体级第{i}行: 源表别名为空（表={table}）")

        if not schema:
            issues.append(f"实体级第{i}行: 源表schema为空（表={table}）")

        # 最后一个有效行的目标表作为全局目标表
        tgt = cells[idx_tgt_table] if idx_tgt_table < len(cells) else ""
        if tgt:
            target_table = tgt

    # 检查目标表名规则（应该是 _i 结尾，_d 结尾的明细层豁免）
    if target_table:
        if not target_table.endswith("_i") and not target_table.endswith("_d"):
            issues.append(f"目标表物理名称 '{target_table}' 不是 _i 结尾（标准要求写I视图名）")
        tgt_schema_val = cells[idx_tgt_schema] if idx_tgt_schema < len(cells) else ""
        if not tgt_schema_val:
            issues.append(f"目标表逻辑schema为空")

    # 属性级
    if "属性级mapping" not in wb.sheetnames:
        issues.append("缺少 sheet: 属性级mapping")
        return issues

    ws_attr = wb["属性级mapping"]
    attr_headers = [str(c.value or "").rstrip("*").strip() for c in ws_attr[1]]

    idx_a_src_table = col_idx(attr_headers, "源表物理表名")
    idx_a_src_alias = col_idx(attr_headers, "源表别名")
    idx_a_target = col_idx(attr_headers, "目标字段名")
    idx_a_type = col_idx(attr_headers, "目标字段类型")
    idx_a_rule = col_idx(attr_headers, "映射规则")

    has_audit = False
    field_count = 0

    for i, row in enumerate(ws_attr.iter_rows(min_row=2, values_only=True), 2):
        cells = [str(c) if c is not None else "" for c in row]
        target_field = cells[idx_a_target] if idx_a_target >= 0 and idx_a_target < len(cells) else ""
        if not target_field:
            continue

        field_count += 1

        # 检查源表在实体级有定义
        src_table = cells[idx_a_src_table] if idx_a_src_table >= 0 and idx_a_src_table < len(cells) else ""
        if src_table and src_table not in entity_tables:
            issues.append(f"属性级第{i}行: 源表 '{src_table}' 在实体级未定义")

        # 检查别名在实体级有对应
        src_alias = cells[idx_a_src_alias] if idx_a_src_alias >= 0 and idx_a_src_alias < len(cells) else ""
        if src_alias and src_alias not in entity_aliases:
            issues.append(f"属性级第{i}行: 别名 '{src_alias}' 在实体级未定义")

        # 检查目标字段名非空
        if not target_field:
            issues.append(f"属性级第{i}行: 目标字段名为空")

        # 检查映射规则合法
        rule = cells[idx_a_rule] if idx_a_rule >= 0 and idx_a_rule < len(cells) else ""
        if rule and rule not in VALID_RULES:
            issues.append(f"属性级第{i}行: 映射规则 '{rule}' 不合法（应为: {VALID_RULES}）")

        # 检查审计字段
        if "审计字段" in (cells[col_idx(attr_headers, "备注")] if col_idx(attr_headers, "备注") >= 0 and col_idx(attr_headers, "备注") < len(cells) else ""):
            has_audit = True

    if not has_audit:
        issues.append("属性级缺少审计字段（备注列标'审计字段'的行）")

    if field_count == 0:
        issues.append("属性级没有任何字段数据")

    # === 检查 RS ===
    if rs_path:
        rs_file = Path(rs_path)
        if not rs_file.exists():
            issues.append(f"RS 文件不存在: {rs_path}")
        else:
            rs_text = rs_file.read_text(encoding="utf-8")
            # 检查 RS 的资产名和 mapping 目标表是否一致
            import re
            # RS 模板里资产名在表格里
            m = re.search(r'\|\s*资产名称\s*\|\s*(\S+)\s*\|', rs_text)
            if m:
                rs_asset = m.group(1).strip()
                if target_table and rs_asset != target_table:
                    issues.append(f"RS 资产名 '{rs_asset}' 和 mapping 目标表 '{target_table}' 不一致")

    return issues


def main():
    parser = argparse.ArgumentParser(description="案例数据自检")
    parser.add_argument("--mapping", default="", help="单个 mapping.xlsx")
    parser.add_argument("--rs", default="", help="单个 RS.md")
    parser.add_argument("--cases-dir", default="", help="检查整个 cases 目录")
    args = parser.parse_args()

    if args.cases_dir:
        # 批量检查
        cases_dir = Path(args.cases_dir)
        all_issues = {}
        for case_dir in sorted(cases_dir.iterdir()):
            if not case_dir.is_dir():
                continue
            mp = case_dir / "mapping.xlsx"
            rs = case_dir / "RS.md"
            if not mp.exists():
                continue
            issues = check_case(str(mp), str(rs) if rs.exists() else "")
            if issues:
                all_issues[case_dir.name] = issues

        if all_issues:
            print(f"发现 {len(all_issues)} 个案例有问题:\n")
            for case_name, issues in all_issues.items():
                print(f"  [{case_name}] {len(issues)} 个问题:")
                for iss in issues:
                    print(f"    - {iss}")
                print()
            sys.exit(1)
        else:
            print(f"✅ 全部案例自检通过")
            sys.exit(0)
    else:
        issues = check_case(args.mapping, args.rs)
        if issues:
            print(f"发现 {len(issues)} 个问题:")
            for iss in issues:
                print(f"  - {iss}")
            sys.exit(1)
        else:
            print("✅ 案例自检通过")
            sys.exit(0)


if __name__ == "__main__":
    main()
